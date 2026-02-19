from __future__ import annotations

import re
import time
from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd


def preco_format(preco_str: Optional[str]) -> float:
    """
    Converte strings como 'R$ 1.200.000' ou '1.200.000,50' para float.
    Retorna 0.0 se nao conseguir converter.
    """
    if preco_str is None:
        return 0.0
    if isinstance(preco_str, (int, float)):
        try:
            return float(preco_str)
        except Exception:
            return 0.0

    s = str(preco_str).strip()
    if not s:
        return 0.0

    s = s.replace("\xa0", " ").strip()

    # Extrai o primeiro numero relevante da string (ex.: "287.000 Cond. R$ 613")
    # para evitar concatenar IPTU/condominio no valor de oferta.
    m = re.search(r"-?\d[\d\.\,]*", s)
    if not m:
        return 0.0

    token = m.group(0)
    token = token.replace(" ", "")
    if not token:
        return 0.0

    # BR com milhar + decimal: 1.234.567,89
    if "," in token and "." in token:
        normalized = token.replace(".", "").replace(",", ".")
    # Apenas virgula: 1234,56
    elif "," in token:
        parts = token.split(",")
        if len(parts) == 2 and len(parts[1]) == 2:
            normalized = token.replace(".", "").replace(",", ".")
        else:
            normalized = token.replace(",", "")
    # Apenas ponto: pode ser milhar (1.355.000) ou decimal (1234.56)
    elif "." in token:
        dot_parts = token.split(".")
        if len(dot_parts) > 1 and all(part.isdigit() for part in dot_parts):
            # Se todos os grupos apos o primeiro tem 3 digitos, trata como milhar.
            if all(len(part) == 3 for part in dot_parts[1:]):
                normalized = "".join(dot_parts)
            else:
                normalized = token
        else:
            normalized = token
    else:
        normalized = token

    try:
        return float(normalized)
    except Exception:
        return 0.0


def aplica_formato_monetario_excel(path: Path, column_name: str = "Valor Oferta (R$)") -> None:
    """Aplica formato monetario (R$) na coluna informada do XLSX."""
    if path.suffix.lower() != ".xlsx":
        return
    try:
        from openpyxl import load_workbook
    except Exception:
        return

    try:
        wb = load_workbook(path)
        ws = wb.active
        col_idx = None
        for cell in ws[1]:
            if cell.value == column_name:
                col_idx = cell.column
                break
        if not col_idx:
            wb.close()
            return

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None or cell.value == "":
                continue
            if isinstance(cell.value, str):
                if not re.search(r"\d", cell.value):
                    continue
                cell.value = preco_format(cell.value)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"R$" #,##0.00'

        wb.save(path)
        wb.close()
    except Exception:
        return


def _excel_points_from_pixels(px_height: int) -> float:
    return max(18.0, min(220.0, round(px_height * 0.75, 2)))


def _header_index_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        value = str(cell.value).strip() if cell.value is not None else ""
        if value:
            headers[value] = cell.column
    return headers


def _ensure_target_column_after(ws, target_column: str, after_column: str) -> int:
    headers = _header_index_map(ws)
    existing = headers.get(target_column)
    if existing:
        return existing

    after_col = headers.get(after_column)
    if after_col:
        target_col = after_col + 1
        ws.insert_cols(target_col, 1)
        ws.cell(row=1, column=target_col, value=target_column)
        return target_col

    target_col = ws.max_column + 1
    ws.cell(row=1, column=target_col, value=target_column)
    return target_col


def _normalize_url(url: str) -> str:
    cleaned = str(url or "").strip()
    if not cleaned:
        return ""
    if cleaned.startswith("//"):
        return f"https:{cleaned}"
    if cleaned.startswith("/"):
        return f"https://www.vivareal.com.br{cleaned}"
    return cleaned if cleaned.startswith("http") else ""


def _extract_og_image_from_listing(listing_url: str, session, timeout: float) -> str:
    listing_url = _normalize_url(listing_url)
    if not listing_url:
        return ""
    try:
        resp = session.get(listing_url, timeout=timeout)
    except Exception:
        return ""
    if not resp.ok:
        return ""

    html = resp.text or ""
    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, flags=re.IGNORECASE)
        if match:
            return _normalize_url(match.group(1))
    return ""


def embute_imagens_excel(
    path: Path,
    image_url_column: str = "Imagem URL",
    target_column: str = "Imagem",
    drop_image_url_column: bool = True,
    max_images: int = 120,
    max_width_px: int = 220,
    max_height_px: int = 150,
    timeout: float = 10.0,
    log_cb=None,
) -> int:
    """
    Faz download das imagens e embute no XLSX para exibicao direta na planilha.
    """
    if path.suffix.lower() != ".xlsx":
        return 0

    try:
        import requests
        from PIL import Image as PILImage
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
    except Exception:
        if log_cb:
            log_cb("Dependencias de imagem ausentes (pillow/requests/openpyxl).")
        return 0

    wb = load_workbook(path)
    ws = wb.active

    target_col = _ensure_target_column_after(ws, target_column=target_column, after_column="UF")
    headers = _header_index_map(ws)
    source_col = headers.get(image_url_column)
    link_col = headers.get("Link Amostra")
    if not source_col and not link_col:
        wb.save(path)
        wb.close()
        if log_cb:
            log_cb(f"Sem '{image_url_column}' e sem 'Link Amostra'. Exportando sem imagens.")
        return 0

    ws.column_dimensions[get_column_letter(target_col)].width = 34

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})

    inserted = 0
    image_buffers: list[BytesIO] = []

    for row_idx in range(2, ws.max_row + 1):
        if max_images and inserted >= max_images:
            break

        raw_image_url = ws.cell(row=row_idx, column=source_col).value if source_col else ""
        image_url = _normalize_url(raw_image_url)

        if not image_url and link_col:
            listing_url = ws.cell(row=row_idx, column=link_col).value
            image_url = _extract_og_image_from_listing(str(listing_url or ""), session, timeout)
            if image_url and source_col:
                ws.cell(row=row_idx, column=source_col, value=image_url)

        if not image_url:
            continue

        try:
            resp = session.get(image_url, timeout=timeout)
            if not resp.ok:
                continue

            content_type = (resp.headers.get("Content-Type") or "").lower()
            if content_type and "image" not in content_type:
                continue

            pil_img = PILImage.open(BytesIO(resp.content))
            if pil_img.mode not in ("RGB", "RGBA"):
                pil_img = pil_img.convert("RGB")
            pil_img.thumbnail((max_width_px, max_height_px))

            png_buffer = BytesIO()
            pil_img.save(png_buffer, format="PNG")
            png_buffer.seek(0)
            image_buffers.append(png_buffer)

            excel_img = XLImage(png_buffer)
            excel_img.width = pil_img.size[0]
            excel_img.height = pil_img.size[1]
            ws.add_image(excel_img, f"{get_column_letter(target_col)}{row_idx}")

            current_height = float(ws.row_dimensions[row_idx].height or 0.0)
            ws.row_dimensions[row_idx].height = max(current_height, _excel_points_from_pixels(int(pil_img.size[1])))
            inserted += 1
        except Exception:
            continue

    if drop_image_url_column and source_col:
        ws.delete_cols(source_col, 1)

    wb.save(path)
    wb.close()

    if log_cb:
        log_cb(f"{inserted} imagens embutidas no XLSX.")
        if max_images and inserted >= max_images:
            log_cb(f"Limite de {max_images} imagens atingido.")

    return inserted


def salva_arquivo(
    df: pd.DataFrame,
    excel_path: Path,
    log_cb=None,
    salvar_csv: bool = False,
    embutir_imagens: bool = False,
    max_imagens: int = 120,
) -> Path:
    """
    Salva DataFrame em Excel (e opcionalmente CSV). Trata arquivo ocupado.
    """
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    path_to_save = excel_path

    if path_to_save.exists():
        try:
            path_to_save.unlink()
        except PermissionError:
            timestamped = path_to_save.with_name(f"{path_to_save.stem}_{int(time.time())}{path_to_save.suffix}")
            if log_cb:
                log_cb("Arquivo estava aberto. Salvando com outro nome...")
            path_to_save = timestamped

    try:
        df.to_excel(path_to_save, index=False)
        aplica_formato_monetario_excel(path_to_save, "Valor Oferta (R$)")
        if embutir_imagens:
            embute_imagens_excel(path_to_save, max_images=max_imagens, log_cb=log_cb)
        if log_cb:
            log_cb(f"{len(df)} linhas salvas em {path_to_save}")
    except PermissionError:
        alt = path_to_save.with_name(f"{path_to_save.stem}_{int(time.time())}{path_to_save.suffix}")
        if log_cb:
            log_cb("Arquivo bloqueado. Tentando salvar com outro nome...")
        df.to_excel(alt, index=False)
        aplica_formato_monetario_excel(alt, "Valor Oferta (R$)")
        if embutir_imagens:
            embute_imagens_excel(alt, max_images=max_imagens, log_cb=log_cb)
        path_to_save = alt
    except ModuleNotFoundError as exc:
        if "openpyxl" in str(exc).lower():
            fallback_csv = path_to_save.with_suffix(".csv")
            try:
                df.to_csv(fallback_csv, index=False, sep=";", encoding="utf-8-sig")
            except PermissionError:
                alt = fallback_csv.with_name(f"{fallback_csv.stem}_{int(time.time())}{fallback_csv.suffix}")
                if log_cb:
                    log_cb("Arquivo CSV bloqueado. Salvando com outro nome...")
                df.to_csv(alt, index=False, sep=";", encoding="utf-8-sig")
                fallback_csv = alt
            path_to_save = fallback_csv
            if log_cb:
                log_cb("Biblioteca 'openpyxl' nao encontrada. Arquivo salvo em CSV como alternativa.")
                log_cb("Instale com 'pip install openpyxl' dentro do ambiente virtual para gerar XLSX.")
        else:
            raise

    if salvar_csv and path_to_save.suffix.lower() != ".csv":
        csv_path = path_to_save.with_suffix(".csv")
        try:
            df.to_csv(csv_path, index=False, sep=";", encoding="utf-8-sig")
        except PermissionError:
            csv_path = csv_path.with_name(f"{csv_path.stem}_{int(time.time())}{csv_path.suffix}")
            if log_cb:
                log_cb("CSV bloqueado. Salvando com outro nome...")
            df.to_csv(csv_path, index=False, sep=";", encoding="utf-8-sig")
        if log_cb:
            log_cb(f"CSV salvo em {csv_path}")

    return path_to_save

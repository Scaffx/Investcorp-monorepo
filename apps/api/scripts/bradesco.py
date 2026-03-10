import io
import re
import unicodedata
import numbers
import logging
from collections import defaultdict
from difflib import get_close_matches
from typing import Iterable

import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side

# Configuração de Logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
log = logging.getLogger(__name__)

# ==========================================
# 1. CONSTANTES E MAPEAMENTOS
# ==========================================

FINAL_ORDER_BRADESCO = [
    "LINHA (NSEQ)", "PROJETO", "BANDEIRA", "CHAVE", "ONDA", "DATA LOTE CLIENTE", "DATA LOTE INVEST",
    "NOME LOTE", "DEADLINE CONCLUSAO", "Contrato", "Junção", "DENOMINACAO/ NOME",
    "TIPO DO LOGRADOURO", "ENDERECO", "NUMERO", "COMPLEMENTO", "BAIRRO", "CIDADE", "UF", "CEP",
    "LATITUDE", "LONGITUDE", "ENDEREÇO CLIENTE", "TIPO IMOVEL", "INFORMACAO ADICIONAL DO IMOVEL",
    "TIPOLOGIA HOMO", "M² AREA TERRENO", "M² AREA CONSTRUIDA", "M² AREA TOTAL", "M² AREA VENDA",
    "M² AREA CONSTRUIDA - CONFERIDO", "M² AREA TOTAL - CONFERIDO", "INDICAR M² PARA ANALISE", "EVIDENCIA",
    "M² MÉDIA DA REGIÃO", "M² MÉDA DA REGIÃO - HOMOGENEIZADO", "COMENTÁRIOS (Adensamento, Capilaridade, Vacância)",
    "DATA DA ANÁLISE", "COMPARATIVO IMÓVEL - MERCADO", "NOME LOCADOR", "NOME LOCADOR - COMPLEMENTO",
    "CNPJ/ CPF LOCADOR", "TELEFONE LOCADOR", "E-MAIL LOCADOR", "NOME FAVORECIDO", "NOME FAVORECIDO - COMPLEMENTO",
    "CNPJ/ CPF FAVORECIDO", "TELEFONE FAVORECIDO", "E-MAIL FAVORECIDO", "NOME CONTATO", "TELEFONE", "EMAIL",
    "INICIO CONTRATO", "TERMINO CONTRATO", "DIA REAJUSTE", "MES REAJUSTE", "DATA PROXIMO REAJUSTE",
    "INDICE", "MES REAJUSTE CONFERIDO", "DATA PROXIMO REAJUSTE - CONFERIDO", "INDICAR DATA REAJUSTE", "EVIDENCIA",
    "ALUGUEL DEVIDO", "ALUGUEL REAJUSTADO - ATUAL / FUTURO (2023)", "ALUGUEL CONFERIDO",
    "INDICAR CONFERIDO / REAJUSTADO / DEVIDO", "EVIDENCIA", "ALUGUEL BASELINE", "% DESCONTO",
    "VALOR DA PROPOSTA", "VALOR POR M² - DEVIDO", "VALOR POR M² - REAJUSTADO", "VALOR POR M² - CONFERIDO",
    "VALOR POR M² - PROPOSTA", "PROPOSTA 1", "CONTRAPROPOSTA", "PROPOSTA 2", "CONTRAPROPOSTA", "STATUS",
    "HISTORICO", "DATA ULTIMO CONTATO", "DATA PROXIMO RETORNO", "MOTIVADORES SEM EXITO", "DATA INICIO NEGOCIACAO",
    "DATA FIM NEGOCIACAO", "DATA SOLICITACAO MINUTA", "DATA ALTERACAO MINUTA", "DATA RECEBIMENTO MINUTA",
    "DATA ENVIO AO LOCADOR", "DATA ENTREGA FORMALIZADA", "RENOVACAO?", "NOVO FIM DE VIGÊNCIA", "PRAZO RENOVADO",
    "REDUCAO?", "VALOR DA REDUÇÃO", "NOVO VALOR DO ALUGUEL", "% REDUCAO", "DATA INICIO CAPTURA", "DATA FIM CAPTURA",
    "PERIODO DE IMPACTO", "DESCONTO?", "VALOR DO DESCONTO (1º PERÍODO)", "NOVO VALOR DO ALUGUEL", "% DESCONTO (1º PERÍODO)",
    "DATA INICIO CAPTURA (1º PERÍODO)", "DATA FIM CAPTURA (1º PERÍODO)", "PERIODO DE IMPACTO (1º PERÍODO)",
    "VALOR DO DESCONTO (2º PERÍODO)", "NOVO VALOR DO ALUGUEL", "% DESCONTO (2º PERÍODO)", "DATA INICIO CAPTURA (2º PERÍODO)",
    "DATA FIM CAPTURA (2º PERÍODO)", "PERIODO DE IMPACTO (2º PERÍODO)", "VALOR DO DESCONTO (3º PERÍODO)",
    "NOVO VALOR DO ALUGUEL", "% DESCONTO (3º PERÍODO)", "DATA INICIO CAPTURA (3º PERÍODO)", "DATA FIM CAPTURA (3º PERÍODO)",
    "PERIODO DE IMPACTO (3º PERÍODO)", "ISENCAO INDICE", "% DE DESCONTO SOBRE O INDICE DE REAJUSTE", "% DE DESCONTO RETIDO",
    "ISENCAO/PARCIAL DATA", "PERIODO CAPTURA", "HOUVE LIMITADOR?", "% LIMITADOR", "LIMITADOR DATA", "SUBSTITUICAO INDICE BASE",
    "NOVO INDICE DE REAJUSTE", "SUBSTITUICAO INDICE DATA", "ALTEROU MES DE REAJUSTE?", "NOVA DATA BASE REAJUSTE",
    "FOI NEGOCIADO SUBSTITUICAO DO FIADOR?", "DATA DA GARANTIA (DATA DO DOC ASSINADO)", "EXCLUSAO GARANTIA", "SUBSTITUICAO GARANTIA",
    "RETROATIVO", "VALOR DEBITO DE RETROATIVOS", "VALOR NEGOCIADO",
    "Economia até  12 meses - Substituição Índice", "Economia até  12 meses Desconto/Redução",
    "Economia até  12 meses - Isenção Indice", "Economia até  12 meses - Limitador do índice",
    "Economia total 12 meses:", "NEGOCIADOR ATUAL",
]

RELNEG_TO_OUTPUT_COLUMN_MAP: dict[str, list[str]] = {
    "BV": ["EB", "DZ"],
    "CJ": ["DY"],
    "CV": ["CQ"],
    "CW": ["CS"],
    "CX": ["CT"],
    "CY": ["CU"],
    "CZ": ["CV"],
}

RAW_COLUMN_ALIASES: dict[str, list[str]] = {
    "LINHA (NSEQ)": ["NSEQ", "LINHA"],
    "ALUGUEL BASELINE": ["ALUGUEL DEVIDO"],
    "ALUGUEL REAJUSTADO - ATUAL / FUTURO (2023)": ["ALUGUEL REAJUSTADO - ATUAL / FUTURO (2023)", "ALUGUEL REAJUSTADO ATUAL FUTURO"],
    "ENDEREÇO": ["ENDEREÇO CLIENTE"],
    "ENDEREÇO CLIENTE": ["ENDERECO CLIENTE", "ENDEREÇO DO CLIENTE"],
    "EMAIL": ["E-MAIL", "E_MAIL"],
    "NOME CONTATO": ["SOLICITANTE"],
    "TELEFONE": ["TEL SOLICITANTE", "TELEFONE SOLICITANTE"],
    "HISTORICO": ["ULTIMO HISTORICO", "ÚLTIMO HISTORICO"],
    "DATA ULTIMO CONTATO": ["DATA HISTORICO", "DATA HISTÓRICO"],
    "DATA ALTERACAO MINUTA": ["DATA ALTERACAO MINUTA"],
    "DATA ENTREGA FORMALIZADA": ["DATA CONCLUSAO", "DATA CONCLUSÃO"],
    "NOVA DATA BASE REAJUSTE": ["NOVA DATA DE REAJUSTE"],
    "PERIODO DE IMPACTO (1º PERÍODO)": ["PERÍODO - RED / DES / MAJ - PREENCHER SOMENTE EM CASO DE DESCONTO", "PERIODO - RED / DES / MAJ - PREENCHER SOMENTE EM CASO DE DESCONTO"],
    "Economia até  12 meses - Substituição Índice": ["Economia ate  12 meses - Substituicao Indice", "Economia ate 12 meses - Substituicao Indice"],
    "Economia até  12 meses Desconto/Redução": ["Economia ate  12 meses Desconto/Reducao", "Economia ate 12 meses Desconto/Reducao"],
    "Economia até  12 meses - Isenção Indice": ["Economia ate  12 meses - Isencao Indice", "Economia ate 12 meses - Isencao Indice"],
    "Economia até  12 meses - Limitador do índice": ["Economia ate  12 meses - Limitador do indice", "Economia ate 12 meses - Limitador do indice"],
}

# ==========================================
# 2. FUNÇÕES AUXILIARES
# ==========================================

def normalize_key(value: str | numbers.Number | None) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^A-Za-z0-9]", "", text).upper()

COLUMN_ALIASES: dict[str, list[str]] = {
    normalize_key(key): [normalize_key(alias) for alias in aliases]
    for key, aliases in RAW_COLUMN_ALIASES.items()
}

def excel_index_to_letter(index: int) -> str:
    if index <= 0:
        raise ValueError("Excel column index must be positive.")
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

FINAL_ORDER_BY_LETTER: dict[str, str] = {
    excel_index_to_letter(idx): column
    for idx, column in enumerate(FINAL_ORDER_BRADESCO, start=1)
}

def to_clean_string(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, numbers.Number):
        return format(value, "g").strip()
    return str(value).strip()

def extrair_dados_locador(valor: object) -> dict[str, str]:
    if not isinstance(valor, str):
        return {}
    texto = unicodedata.normalize("NFKD", valor)
    texto = texto.replace("\r", "\n")
    resultado: dict[str, str] = {}
    
    email_match = re.search(r"[\w\.-]+@[\w\.-]+", texto)
    if email_match:
        resultado["email"] = email_match.group(0)
        
    doc_match = re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|\d{11}", re.sub(r"[^\d/@.-]", "", texto))
    if doc_match:
        resultado["documento"] = doc_match.group(0)
        
    tel_match = re.search(r"\(?\d{2}\)?\s?\d{4,5}-?\d{4}", texto)
    if tel_match:
        resultado["telefone"] = tel_match.group(0)
        
    linhas = [ln.strip() for ln in texto.splitlines() if ln.strip()]
    if linhas:
        resultado["nome"] = linhas[0]
        
    return resultado

def _best_column_match(normalized_columns: dict[str, list[str]], target: str) -> str | None:
    keys = [key for key, cols in normalized_columns.items() if cols]
    if not keys:
        return None
    match = get_close_matches(target, keys, n=1, cutoff=0.9)
    if match:
        return normalized_columns[match[0]].pop(0)
    return None

def _normalize_mapping(columns: Iterable[str]) -> dict[str, list[str]]:
    mapping: defaultdict[str, list[str]] = defaultdict(list)
    for column in columns:
        mapping[normalize_key(column)].append(column)
    return mapping

def parse_currency_to_float(value: object) -> float:
    text = to_clean_string(value).replace("R$", "").strip()
    if not text:
        return 0.0
    normalized = text.replace(".", "").replace(",", ".").replace(" ", "")
    try:
        return float(normalized)
    except ValueError:
        digits = re.findall(r"-?\d+(?:[.,]\d+)?", text)
        if digits:
            try:
                return float(digits[0].replace(",", "."))
            except ValueError:
                return 0.0
    return 0.0

def format_currency_value(value: object) -> str:
    text = to_clean_string(value)
    if not text:
        return ""
    cleaned = text.replace("R$", "").strip()
    return f"R$ {cleaned}" if cleaned else "R$"

# ==========================================
# 3. FUNÇÃO PRINCIPAL (MOTOR DE DADOS)
# ==========================================

def processar_relatorio_bradesco(arquivo_excel_em_memoria, nseq_string: str) -> io.BytesIO:
    """
    Processa a planilha de renegociação filtrando pelos NSEQs informados,
    aplica as regras de negócio e retorna um buffer de memória com o Excel formatado.
    """
    # 1. Preparar os NSEQs recebidos do frontend (substitui a leitura do TXT)
    nseq_ordenados = [str(n).strip() for n in nseq_string.split(',') if str(n).strip()]
    if not nseq_ordenados:
        raise ValueError("Nenhum NSEQ foi informado para gerar o relatório.")

    # Simula a ordem e a linha_regra que antes vinham do TXT
    ordem_map = {nseq: idx for idx, nseq in enumerate(nseq_ordenados)}
    linha_por_nseq = {nseq: str(idx + 1) for idx, nseq in enumerate(nseq_ordenados)}

    log.info("Total de NSEQ recebidos: %s", len(nseq_ordenados))

    # 2. Ler o arquivo diretamente da memória
    df_neg = pd.read_excel(arquivo_excel_em_memoria, dtype=str, engine="openpyxl").fillna("")
    df_neg.columns = df_neg.columns.str.strip().str.lower()

    column_letter_map = {
        excel_index_to_letter(idx): column
        for idx, column in enumerate(df_neg.columns, start=1)
    }

    if "nseq" not in df_neg.columns:
        raise ValueError("Coluna 'nseq' não encontrada na planilha anexada.")

    if "centro de custo" not in df_neg.columns and "chave" not in df_neg.columns:
        raise ValueError("Coluna de chave (centro de custo ou chave) não encontrada.")

    if "chave" not in df_neg.columns:
        df_neg = df_neg.rename(columns={"centro de custo": "chave"})

    df_neg["chave"] = df_neg["chave"].astype(str).str.strip().str.upper()
    df_neg["nseq"] = df_neg["nseq"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

    # 3. Filtrar e Ordenar
    nseq_lookup = set(nseq_ordenados)
    df_filtrado = df_neg[df_neg["nseq"].isin(nseq_lookup)].copy()

    if df_filtrado.empty:
        raise ValueError("Nenhum dos NSEQs informados foi encontrado na planilha anexada.")

    df_filtrado["ordem"] = df_filtrado["nseq"].map(ordem_map)
    df_filtrado = df_filtrado.dropna(subset=["ordem"]).sort_values(by="ordem")

    df_filtrado["linha_regra"] = df_filtrado["nseq"].map(linha_por_nseq).fillna("")
    df_filtrado["linha_regra"] = df_filtrado["linha_regra"].apply(
        lambda x: int(x) if isinstance(x, str) and x.isdigit() else x
    )

    preenchimento = df_neg.replace("", pd.NA).notna().sum(axis=1)
    df_neg["__row_quality"] = preenchimento

    df_base = (
        df_neg.sort_values(["chave", "__row_quality"], ascending=[True, False])
        .drop_duplicates(subset="chave", keep="first")
        .drop(columns="__row_quality", errors="ignore")
    )

    df_base_sem_nseq = df_base.drop(columns=["nseq"], errors="ignore")

    df_merged = pd.merge(
        df_filtrado[["chave", "linha_regra", "nseq"]],
        df_base_sem_nseq,
        on="chave",
        how="left",
    )

    # 4. Mapeamento de Colunas
    normalized_columns = _normalize_mapping(df_merged.columns)
    ordered_series = []
    missing_columns = []

    for target_col in FINAL_ORDER_BRADESCO:
        norm_target = normalize_key(target_col)
        candidate = None
        candidates = COLUMN_ALIASES.get(norm_target, [])
        search_keys = [norm_target, *candidates]

        for key in search_keys:
            if key in normalized_columns and normalized_columns[key]:
                candidate = normalized_columns[key].pop(0)
                break

        if candidate is None:
            candidate = _best_column_match(normalized_columns, norm_target)

        if candidate:
            series = df_merged[candidate].copy()
            series.name = target_col
        else:
            series = pd.Series("", index=df_merged.index, name=target_col)
            missing_columns.append(target_col)

        ordered_series.append(series)

    if missing_columns:
        log.debug("Colunas não encontradas: %s", missing_columns)

    df_ordenado = pd.concat(ordered_series, axis=1)

    # 5. Preenchimento e Tratamento de Dados
    def preencher_se_vazio(destino: str, origem: str) -> None:
        if destino in df_ordenado.columns and origem in df_merged.columns:
            df_ordenado[destino] = df_ordenado[destino].mask(
                df_ordenado[destino] == "", df_merged[origem]
            )

    preencher_se_vazio("NOME CONTATO", "solicitante")
    preencher_se_vazio("TELEFONE", "tel solicitante")
    preencher_se_vazio("EMAIL", "e-mail solicitante")
    preencher_se_vazio("NEGOCIADOR ATUAL", "negociador")
    preencher_se_vazio("LINHA (NSEQ)", "nseq")
    preencher_se_vazio("LINHA (NSEQ)", "linha_regra")
    preencher_se_vazio("PROJETO", "empresa")
    preencher_se_vazio("ALUGUEL BASELINE", "aluguel devido")
    preencher_se_vazio("DATA ULTIMO CONTATO", "data historico")
    preencher_se_vazio("DATA ALTERACAO MINUTA", "data alteracao minuta")
    preencher_se_vazio("DATA ENTREGA FORMALIZADA", "data conclusão")
    preencher_se_vazio("PERIODO DE IMPACTO (1º PERÍODO)", "período - red / des / maj - preencher somente em caso de desconto")

    economia_cols = [
        "Economia até  12 meses - Substituição Índice",
        "Economia até  12 meses Desconto/Redução",
        "Economia até  12 meses - Isenção Indice",
        "Economia até  12 meses - Limitador do índice",
    ]
    total_economia_col = "Economia total 12 meses:"

    if all(col in df_ordenado.columns for col in economia_cols):
        def calcular_total_economia(row: pd.Series) -> str:
            valores = [parse_currency_to_float(row[col]) for col in economia_cols]
            total = sum(valores)
            if total == 0:
                return ""
            return f"{total:.2f}"
        df_ordenado[total_economia_col] = df_ordenado.apply(calcular_total_economia, axis=1)

    for col_currency in [
        "ALUGUEL REAJUSTADO - ATUAL / FUTURO (2023)",
        "ALUGUEL BASELINE",
        "ALUGUEL DEVIDO",
        *economia_cols,
        total_economia_col,
    ]:
        if col_currency in df_ordenado.columns:
            df_ordenado[col_currency] = df_ordenado[col_currency].apply(format_currency_value)

    for source_letter, dest_letters in RELNEG_TO_OUTPUT_COLUMN_MAP.items():
        source_col = column_letter_map.get(source_letter.upper())
        if not source_col or source_col not in df_merged.columns:
            continue
        for dest_letter in dest_letters if isinstance(dest_letters, (list, tuple)) else [dest_letters]:
            dest_col = FINAL_ORDER_BY_LETTER.get(dest_letter.upper())
            if not dest_col or dest_col not in df_ordenado.columns:
                continue
            df_ordenado[dest_col] = df_merged[source_col].reindex(df_ordenado.index).fillna("")

    if "dados locador (a)" in df_merged.columns:
        locador_info = df_merged["dados locador (a)"].apply(extrair_dados_locador)
        for destino, chave_info in {
            "NOME LOCADOR": "nome",
            "CNPJ/ CPF LOCADOR": "documento",
            "TELEFONE LOCADOR": "telefone",
            "E-MAIL LOCADOR": "email",
        }.items():
            if destino in df_ordenado.columns:
                df_ordenado[destino] = df_ordenado[destino].mask(
                    df_ordenado[destino] == "",
                    locador_info.apply(lambda dados: dados.get(chave_info, "")),
                )

    if {"CHAVE", "ONDA", "Contrato"}.issubset(df_ordenado.columns):
        chave_original = df_ordenado["CHAVE"].map(to_clean_string)
        onda_formatada = df_ordenado["ONDA"].map(to_clean_string)
        df_ordenado["Contrato"] = chave_original
        df_ordenado["CHAVE"] = chave_original.combine(
            onda_formatada,
            lambda chave, onda: f"{chave}-{onda}" if chave and onda else chave,
        )

    if "ENDERECO" not in df_ordenado.columns:
        df_ordenado["ENDERECO"] = ""
    if "ENDEREÇO CLIENTE" not in df_ordenado.columns:
        df_ordenado["ENDEREÇO CLIENTE"] = ""

    base_endereco = (
        df_ordenado[["ENDERECO", "ENDEREÇO CLIENTE"]]
        .replace("", pd.NA)
        .bfill(axis=1)
        .iloc[:, 0]
        .fillna("")
    )
    df_ordenado["ENDERECO"] = base_endereco
    df_ordenado["ENDEREÇO CLIENTE"] = base_endereco

    # ==========================================
    # 6. GERAÇÃO DO EXCEL E COLORAÇÃO EM MEMÓRIA
    # ==========================================
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_ordenado.to_excel(writer, index=False, sheet_name='Relatorio')
        
        # Acessando o workbook para aplicar a formatação visual
        workbook = writer.book
        worksheet = writer.sheets['Relatorio']
        
        cinza = PatternFill(start_color="c6c6c6", end_color="c6c6c6", fill_type="solid")
        azul = PatternFill(start_color="00008b", end_color="00008b", fill_type="solid")
        azul_claro = PatternFill(start_color="21abcd", end_color="21abcd", fill_type="solid")
        preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cinza_escuro = PatternFill(start_color="a0a0a0", end_color="a0a0a0", fill_type="solid")

        faixas = [
            (1, 12, cinza), (13, 22, preto), (23, 23, cinza), (24, 26, azul),
            (27, 30, cinza), (31, 34, azul), (35, 36, cinza), (37, 38, azul),
            (39, 49, cinza), (50, 52, azul), (53, 58, cinza), (59, 59, azul),
            (60, 60, cinza), (61, 62, azul), (63, 64, cinza), (65, 67, azul),
            (68, 68, cinza), (69, 69, azul), (70, 74, cinza), (75, 90, azul),
            (91, 92, azul_claro), (93, 93, cinza_escuro), (94, 95, azul_claro), (96, 97, cinza_escuro),
            (98, 98, azul), (99, 100, cinza), (101, 102, azul_claro), (103, 104, cinza_escuro),
            (105, 106, azul_claro), (107, 107, cinza_escuro), (108, 108, azul),
            (109, 110, cinza_escuro), (111, 112, azul_claro), (113, 113, cinza_escuro),
            (114, 114, azul_claro), (115, 116, cinza_escuro), (117, 118, azul_claro),
            (119, 119, cinza_escuro), (120, 121, azul_claro), (122, 122, cinza_escuro),
            (123, 144, azul_claro)
        ]

        for start, end, color in faixas:
            for col in range(start, end + 1):
                worksheet.cell(row=1, column=col).fill = color

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            if cell.value is not None and str(cell.value).strip():
                cell.font = Font(color="FFFFFF", bold=True)

        worksheet.freeze_panes = "M1"
        worksheet.auto_filter.ref = worksheet.dimensions

        thin = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        for r in range(1, worksheet.max_row + 1):
            for c in range(1, worksheet.max_column + 1):
                worksheet.cell(row=r, column=c).border = border_all

        worksheet.sheet_view.showGridLines = True

    output.seek(0)
    return output

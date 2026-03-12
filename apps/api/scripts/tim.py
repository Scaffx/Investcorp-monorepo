#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Optional, Iterable, Union, Sequence
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# ====== CONFIGURAÇÕES DE CAMINHOS PARA SERVIDOR ======
BASE_DIR = Path(__file__).resolve().parent.parent
MODELOS_DIR = BASE_DIR / "modelos"

OUTPUT_PREFIX = "TIM"

DEFAULT_STATUS_COL = "DQ"
DEFAULT_X_COL = "X"
DEFAULT_EB_COL = "EB"

# REGRA 4: Removido o contrato 3006864 desta lista
ALWAYS_INCLUDE_CONTRACTS = {"3002893", "3005931"}
# Contratos que DEVEM ser ignorados, mesmo que venham no SIIM
FORBIDDEN_CONTRACTS = {"3006864"}

HISTORY_HEADER_CANDIDATES = ("ULTIMO HISTORICO", "ÚLTIMO HISTORICO", "ULTIMO HISTÓRICO", "ULTIMO HIST", "ULTIMO HISTORICO ", "EC", "EC (ULTIMO HISTORICO)", "EC ULTIMO HISTORICO")
NSEQ_HEADER_CANDIDATES = ("NSEQ", "NSEQ SIIM", "NSEQ_SIIM", "NSEQ SIIM - TIM", "NSEQ ESCOLHIDO")
ONDA_HEADER_CANDIDATES = ("ONDA",)
CONTRACT_HEADER_CANDIDATES = ("CONTRATO", "CONTRATO SAP", "ORDEM SAP", "ORDEM_SAP", "ORDEM SAP (OC)")

# Novas colunas para buscar no SIIM (Regras 1 e 2)
INICIO_CONTRATO_CANDIDATES = ("INICIO CONTRATO", "DATA INICIO", "DATA_INICIO")
FIM_CONTRATO_CANDIDATES = ("TERMINO CONTRATO", "DATA FIM", "DATA_FIM")

MODEL_SHEET_FALLBACKS = ("INVESTCORP", "BASE", "Planilha1", "Dados", "TIM")

# ====== FUNÇÕES AUXILIARES ======

def _find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "CONTRATO":
                return cell.row
    return 1

def excel_col_to_idx(col_letter: str) -> int:
    col_letter = col_letter.strip().upper()
    result = 0
    for ch in col_letter:
        if not ("A" <= ch <= "Z"): raise ValueError(f"Invalid column letter: {col_letter}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result

def ensure_width(df: pd.DataFrame, min_cols: int) -> pd.DataFrame:
    need = max(0, min_cols - df.shape[1])
    if need > 0:
        for _ in range(need): df[df.shape[1]] = pd.NA
    return df

def tail_after_last_hyphen(value: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)): return ""
    text = str(value)
    parts = text.split(" - ") if " - " in text else text.split("-")
    return parts[-1].strip() if parts else text.strip()

def _normalize_nseq_value(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)): return ""
    text = str(value).strip()
    if not text: return ""
    if text.endswith(".0"): text = text[:-2]
    return text

def _find_column_index_by_label(df: pd.DataFrame, candidates: Sequence[str], max_rows: int = 50) -> Optional[int]:
    if df.empty: return None
    target_labels = {candidate.strip().upper() for candidate in candidates}
    rows_to_scan = df.head(max_rows)
    for _, row in rows_to_scan.iterrows():
        for idx, raw in enumerate(row):
            value = str(raw).strip().upper()
            if value in target_labels: return idx
    return None

def _normalize_label(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)): return ""
    return str(value).replace("\n", " ").strip()

def _build_counts(series: pd.Series, empty_label: Optional[str] = None) -> dict[str, int]:
    from collections import Counter
    normalized: list[str] = []
    for value in series:
        label = _normalize_label(value)
        if not label:
            if empty_label is None: continue
            label = empty_label
        normalized.append(label.upper())
    counts = Counter(normalized)
    return {key: int(val) for key, val in counts.items()}

def _apply_counts_to_table(ws, header_row: int, counts: dict[str, int], total: int) -> None:
    row = header_row + 1
    while True:
        label = ws.cell(row=row, column=1).value
        normalized = _normalize_label(label)
        if not normalized: break
        key = normalized.upper()
        if key.startswith("TOTAL GERAL"):
            ws.cell(row=row, column=2, value=total)
            break
        ws.cell(row=row, column=2, value=int(counts.get(key, 0)))
        row += 1

def _normalize_onda_value(value: object) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)): return 0
    try: return int(float(str(value).strip()))
    except Exception: return 0

def _resolve_sheet_name(sheet: Optional[Union[str, int]], workbook_file, fallback_names: Sequence[str] = ()) -> Union[str, int]:
    if sheet is None: desired = None
    elif isinstance(sheet, str): desired = sheet.strip()
    else: desired = sheet
    if desired not in (None, ""): return desired
    
    if workbook_file is not None and fallback_names:
        try:
            if hasattr(workbook_file, 'seek'):
                workbook_file.seek(0)
            xls = pd.ExcelFile(workbook_file)
            for candidate in fallback_names:
                if candidate in xls.sheet_names: return candidate
        except Exception: pass
    return 0

# REGRA 1 e 2: Formatação de Data DD/MM/AA
def format_date_ddmmaa(value: object) -> str:
    if pd.isna(value) or value is None or str(value).strip() == "":
        return ""
    try:
        # Tenta converter para datetime
        dt = pd.to_datetime(value, errors='coerce', dayfirst=True)
        if pd.isna(dt):
            return str(value)
        # Formata como DD/MM/AA
        return dt.strftime("%d/%m/%y")
    except Exception:
        return str(value)

# REGRA 3: Limpar "TIM - " do Status
def clean_tim_status(value: object) -> str:
    if pd.isna(value) or value is None:
        return ""
    text = str(value).strip()
    # Remove o prefixo "TIM - " (case insensitive)
    return re.sub(r'(?i)^TIM\s*-\s*', '', text)

# ====== CONSTRUÇÃO DO DATAFRAME ======

def _build_tim_dataframe(modelo_file, relneg_file, allowed_nseqs: Iterable[str]) -> pd.DataFrame:
    allowed_nseqs_norm = {_normalize_nseq_value(item) for item in allowed_nseqs if _normalize_nseq_value(item)}
    
    modelo_sheet_name = _resolve_sheet_name(None, workbook_file=modelo_file, fallback_names=MODEL_SHEET_FALLBACKS)
    
    if hasattr(modelo_file, 'seek'):
        modelo_file.seek(0)
    modelo = pd.read_excel(modelo_file, sheet_name=modelo_sheet_name, header=None, dtype=object)
    
    needed_cols_modelo = max(excel_col_to_idx("U"), excel_col_to_idx("W"), excel_col_to_idx(DEFAULT_X_COL))
    modelo = ensure_width(modelo, needed_cols_modelo)
    a_to_u = modelo.iloc[:, : excel_col_to_idx("U")]
    
    model_nseq_idx = _find_column_index_by_label(a_to_u, NSEQ_HEADER_CANDIDATES)
    model_contract_idx = _find_column_index_by_label(modelo, CONTRACT_HEADER_CANDIDATES)
    if model_contract_idx is None: raise ValueError("Coluna de contrato não encontrada no modelo TIM.")
    
    w_series = modelo.iloc[:, excel_col_to_idx("W") - 1]
    
    if hasattr(relneg_file, 'seek'):
        relneg_file.seek(0)
    rel = pd.read_excel(relneg_file, header=None, dtype=object)
    
    needed_cols_rel = max(excel_col_to_idx(DEFAULT_EB_COL), excel_col_to_idx(DEFAULT_STATUS_COL))
    rel = ensure_width(rel, needed_cols_rel)
    
    nseq_idx = _find_column_index_by_label(rel, NSEQ_HEADER_CANDIDATES)
    if nseq_idx is None: raise ValueError("Coluna NSEQ não encontrada no RelNegociacao.")
    
    contract_idx = _find_column_index_by_label(rel, CONTRACT_HEADER_CANDIDATES)
    if contract_idx is None: raise ValueError("Coluna de contrato não encontrada no RelNegociacao.")
    
    history_idx = _find_column_index_by_label(rel, HISTORY_HEADER_CANDIDATES)
    inicio_idx = _find_column_index_by_label(rel, INICIO_CONTRATO_CANDIDATES)
    fim_idx = _find_column_index_by_label(rel, FIM_CONTRATO_CANDIDATES)
    
    eb_idx = excel_col_to_idx(DEFAULT_EB_COL) - 1
    dq_idx = excel_col_to_idx(DEFAULT_STATUS_COL) - 1
    
    rel_df = pd.DataFrame({
        "EB_raw": rel.iloc[:, eb_idx],
        "EB": rel.iloc[:, eb_idx].astype(str).str.strip(),
        "DQ": rel.iloc[:, dq_idx],
        "CONTRATO_norm": rel.iloc[:, contract_idx].apply(lambda v: "" if pd.isna(v) else str(v).strip()),
    })
    
    rel_df = rel_df.loc[rel_df["CONTRATO_norm"] != ""].copy()
    
    # REGRA 4: Remover explicitamente o contrato proibido
    rel_df = rel_df.loc[~rel_df["CONTRATO_norm"].isin(FORBIDDEN_CONTRACTS)].copy()
    
    rel_df["HISTORY_raw"] = rel.iloc[:, history_idx] if history_idx is not None else pd.NA
    rel_df["INICIO_raw"] = rel.iloc[:, inicio_idx] if inicio_idx is not None else pd.NA
    rel_df["FIM_raw"] = rel.iloc[:, fim_idx] if fim_idx is not None else pd.NA
    
    rel_df["NSEQ_norm"] = rel.iloc[:, nseq_idx].apply(_normalize_nseq_value)
    
    if allowed_nseqs_norm:
        rel_df = rel_df.loc[rel_df["NSEQ_norm"].isin(allowed_nseqs_norm)].copy()
        
    onda_idx = _find_column_index_by_label(rel, ONDA_HEADER_CANDIDATES)
    rel_df["ONDA_val"] = rel.iloc[:, onda_idx].apply(_normalize_onda_value) if onda_idx is not None else 0
    
    rel_df = rel_df.sort_values(by=["NSEQ_norm", "ONDA_val"], ascending=[True, False]).drop_duplicates(subset=["NSEQ_norm"], keep="first")
    
    # REGRA 3: Aplicar limpeza do status "TIM - "
    rel_df["V_status_tail"] = rel_df["DQ"].apply(clean_tim_status)
    
    # REGRA 1 e 2: Formatar datas
    rel_df["INICIO_fmt"] = rel_df["INICIO_raw"].apply(format_date_ddmmaa)
    rel_df["FIM_fmt"] = rel_df["FIM_raw"].apply(format_date_ddmmaa)
    
    contract_status_map = rel_df.dropna(subset=["CONTRATO_norm"]).drop_duplicates(subset=["CONTRATO_norm"], keep="first").set_index("CONTRATO_norm")["V_status_tail"].to_dict()
    contract_to_eb_raw = rel_df.dropna(subset=["CONTRATO_norm"]).drop_duplicates(subset=["CONTRATO_norm"], keep="first").set_index("CONTRATO_norm")["EB_raw"].to_dict()
    contract_history_map = rel_df.dropna(subset=["CONTRATO_norm"]).drop_duplicates(subset=["CONTRATO_norm"], keep="first").set_index("CONTRATO_norm")["HISTORY_raw"].to_dict() if "HISTORY_raw" in rel_df.columns else {}
    
    contract_inicio_map = rel_df.dropna(subset=["CONTRATO_norm"]).drop_duplicates(subset=["CONTRATO_norm"], keep="first").set_index("CONTRATO_norm")["INICIO_fmt"].to_dict()
    contract_fim_map = rel_df.dropna(subset=["CONTRATO_norm"]).drop_duplicates(subset=["CONTRATO_norm"], keep="first").set_index("CONTRATO_norm")["FIM_fmt"].to_dict()
    
    allowed_contracts = set(contract_status_map.keys())
    contract_series = modelo.iloc[:, model_contract_idx].apply(lambda v: "" if pd.isna(v) else str(v).strip())
    
    contract_to_index = {contract: idx for idx, contract in enumerate(contract_series) if contract}
    
    ordered_contracts = []
    for contract in [c for c in rel_df["CONTRATO_norm"] if c in allowed_contracts]:
        if contract not in ordered_contracts: ordered_contracts.append(contract)
    for contract in ALWAYS_INCLUDE_CONTRACTS:
        if contract and contract not in ordered_contracts: ordered_contracts.append(contract)
        
    rows_a_to_u, v_values, history_values, eb_values = [], [], [], []
    
    # Encontrar as colunas de Vigência no modelo para substituí-las
    header_row_idx = 0
    for idx, row in modelo.iterrows():
        if any(str(val).strip().upper() == "CONTRATO" for val in row if pd.notna(val)):
            header_row_idx = idx
            break
            
    headers = [str(col).strip().upper() for col in modelo.iloc[header_row_idx]]
    
    inicio_col_idx = -1
    fim_col_idx = -1
    
    for idx, h in enumerate(headers):
        if "VIGÊNCIA ATUALIZADA INÍCIO" in h or "VIGENCIA ATUALIZADA INICIO" in h:
            inicio_col_idx = idx
        elif "VIGÊNCIA ATUALIZADA FIM" in h or "VIGENCIA ATUALIZADA FIM" in h:
            fim_col_idx = idx
    
    for contract in ordered_contracts:
        idx_in_model = contract_to_index.get(contract)
        if idx_in_model is not None:
            row_series = a_to_u.iloc[idx_in_model].copy()
            fallback_history = w_series.iloc[idx_in_model]
        else:
            row_series = pd.Series(pd.NA, index=a_to_u.columns)
            row_series.iloc[model_contract_idx] = contract
            fallback_history = pd.NA
            
        # Aplica as datas formatadas (Regras 1 e 2)
        if inicio_col_idx != -1 and inicio_col_idx < len(row_series):
            row_series.iloc[inicio_col_idx] = contract_inicio_map.get(contract, row_series.iloc[inicio_col_idx])
            
        if fim_col_idx != -1 and fim_col_idx < len(row_series):
            row_series.iloc[fim_col_idx] = contract_fim_map.get(contract, row_series.iloc[fim_col_idx])
            
        rows_a_to_u.append(row_series)
        v_values.append(contract_status_map.get(contract, ""))
        
        history_value = contract_history_map.get(contract, fallback_history)
        history_values.append(fallback_history if pd.isna(history_value) else history_value)
        eb_values.append(contract_to_eb_raw.get(contract, ""))
        
    out = pd.DataFrame(rows_a_to_u).reset_index(drop=True)
    out = pd.concat([out, pd.Series(v_values), pd.Series(history_values), pd.Series(eb_values)], axis=1)
    
    if model_nseq_idx is not None and model_nseq_idx < out.shape[1]:
        out = out.drop(out.columns[model_nseq_idx], axis=1)
        
    def idx_to_excel(idx: int) -> str:
        letters = ""
        while idx:
            idx, remainder = divmod(idx - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters
    
    out.columns = [idx_to_excel(i) for i in range(1, out.shape[1] + 1)]
    return out

# ====== NÚCLEO ADAPTADO PARA API ======

def processar_relatorio_tim(arquivo_excel_em_memoria, nseq_string: str, arquivo_modelo_em_memoria=None) -> io.BytesIO:
    """
    Função principal adaptada para receber dados da API em memória.
    """
    if not nseq_string:
        raise ValueError("Nenhum NSEQ fornecido para processamento.")
        
    nseq_list = [n.strip() for n in nseq_string.split(',') if n.strip()]

    # Lógica do Modelo
    if arquivo_modelo_em_memoria:
        modelo_file = arquivo_modelo_em_memoria
    else:
        # Busca no servidor
        candidates = sorted(MODELOS_DIR.glob("TIM*.xlsx"))
        modelo_file = candidates[0] if candidates else MODELOS_DIR / "TIM_Modelo.xlsx"
        if not modelo_file.exists():
            raise FileNotFoundError("Planilha de modelo TIM não encontrada no servidor.")

    # Processa o DataFrame
    df_out = _build_tim_dataframe(modelo_file, arquivo_excel_em_memoria, nseq_list)
    
    # Volta o ponteiro do modelo para carregar o openpyxl
    if hasattr(modelo_file, 'seek'):
        modelo_file.seek(0)
    
    modelo_sheet_name = _resolve_sheet_name(None, workbook_file=modelo_file, fallback_names=MODEL_SHEET_FALLBACKS)
    wb = load_workbook(modelo_file)
    ws = wb[modelo_sheet_name]
    
    header_row = _find_header_row(ws)
    data_start_row = header_row + 1
    
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)
        
    for row_offset, row_values in enumerate(df_out.itertuples(index=False, name=None)):
        target_row = data_start_row + row_offset
        for col_offset, value in enumerate(row_values, start=1):
            ws.cell(row=target_row, column=col_offset, value=None if pd.isna(value) else value)
            
    total_rows = len(df_out)
    
    # REGRA 5: Ajustar colunas B e Q para conter a QTD de contratos
    # Coluna B = 2, Coluna Q = 17
    ws.cell(row=2, column=2, value=total_rows)   # Coluna B (Quantidade)
    ws.cell(row=2, column=17, value=total_rows)  # Coluna Q (Quantidade)
    
    total_cols = df_out.shape[1] if df_out.shape[1] else ws.max_column
    last_row_for_border = data_start_row + total_rows - 1 if total_rows else header_row
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    border_thin = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))
    
    for col_idx in (1, 4):
        if col_idx <= ws.max_column:
            for row in range(header_row, last_row_for_border + 1):
                ws.cell(row=row, column=col_idx).alignment = center_alignment
                
    for row in range(header_row, last_row_for_border + 1):
        for col_idx in range(1, total_cols + 1):
            ws.cell(row=row, column=col_idx).border = border_thin
            
    header_map = {str(ws.cell(header_row, c).value).strip().upper(): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    data_rows = list(range(data_start_row, data_start_row + total_rows))
    
    if "resumo" in wb.sheetnames:
        ws_resumo = wb["resumo"]
        resumo_headers = [r for r in range(1, ws_resumo.max_row + 1) if _normalize_label(ws_resumo.cell(r, 1).value).upper() == "RÓTULOS DE LINHA"]
        if resumo_headers:
            col_vigencia = header_map.get("VIGÊNCIA ATUALIZADA")
            col_tipo = header_map.get("TIPO DE LOCADOR")
            if col_vigencia:
                counts_vigencia = _build_counts(pd.Series([ws.cell(row=r, column=col_vigencia).value for r in data_rows]))
                _apply_counts_to_table(ws_resumo, resumo_headers[0], counts_vigencia, total_rows)
            if len(resumo_headers) > 1 and col_tipo:
                counts_tipo = _build_counts(pd.Series([ws.cell(row=r, column=col_tipo).value for r in data_rows]))
                _apply_counts_to_table(ws_resumo, resumo_headers[1], counts_tipo, total_rows)

    # Salva na memória RAM
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

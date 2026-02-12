#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Generate Claro Renovacao and Distrato reports and combine into a single workbook."""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, timedelta
import sys
from importlib import import_module
from copy import copy

import pandas as pd
from openpyxl import load_workbook, Workbook

try:
    from .utils import show_generation_popup
except ImportError:  # Executando diretamente sem pacote
    current_dir = Path(__file__).resolve().parent
    if str(current_dir) not in sys.path:
        sys.path.append(str(current_dir))
    from utils import show_generation_popup  # type: ignore[import-not-found]

BASE_DIR = Path.home() / "Desktop" / "Report"
RENOVACAO_PREFIX = "CLARO_RENOVACAO"
DISTRATO_PREFIX = "CLARO_DISTRATO"
FINAL_PREFIX = "Claro_Renova"



def _count_rows(path: Path) -> int:
    workbook = pd.read_excel(path, sheet_name=None)
    return sum(len(df) for df in workbook.values())


def _paths_for_yesterday(base_dir: Path) -> tuple[Path, Path, Path, str, str]:
    yesterday = datetime.now() - timedelta(days=1)
    folder_token = yesterday.strftime("%d-%m-%Y")
    file_token = yesterday.strftime("%d-%m-%Y")
    out_dir = base_dir / folder_token
    ren_path = out_dir / f"{RENOVACAO_PREFIX}_{file_token}.xlsx"
    dis_path = out_dir / f"{DISTRATO_PREFIX}_{file_token}.xlsx"
    final_path = out_dir / f"{FINAL_PREFIX}_{file_token}.xlsx"
    return ren_path, dis_path, final_path, folder_token, file_token


def _ensure_report(build_callable, label: str, expected_path: Path) -> tuple[Path, int, Path]:
    """Execute report builder and return (path, rows, output_dir)."""
    try:
        result = build_callable()
    except FileNotFoundError as exc:
        print(f"[WARN] {label}: {exc}")
        if expected_path.exists():
            try:
                rows = _count_rows(expected_path)
            except Exception:
                rows = 0
            return expected_path, rows, expected_path.parent
        raise FileNotFoundError(f"{label}: {exc}") from exc

    if isinstance(result, tuple) and len(result) == 3:
        path, rows, out_dir = result
        if not rows:
            try:
                rows = _count_rows(path)
            except Exception:
                rows = 0
        return path, rows, out_dir

    if isinstance(result, Path):
        path = result
        try:
            rows = _count_rows(path)
        except Exception:
            rows = 0
        return path, rows, path.parent

    raise TypeError(f"{label}: unexpected return from report generator")


def _copy_sheet(source_ws, target_wb: Workbook, title: str) -> None:
    target_ws = target_wb.create_sheet(title=title)

    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)

    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
    for idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx].height = dim.height

    if source_ws.sheet_format.defaultColWidth is not None:
        target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
    if source_ws.sheet_format.defaultRowHeight is not None:
        target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.auto_filter = source_ws.auto_filter


def merge_reports() -> Path:
    ren_expected, dis_expected, final_path, folder_token, _ = _paths_for_yesterday(BASE_DIR)

    def _import_script_module(module_basename: str):
        candidates = []
        if __package__:
            candidates.append(f"{__package__}.{module_basename}")
        candidates.append(f"scripts.{module_basename}")
        candidates.append(module_basename)

        last_error: Exception | None = None
        for candidate in candidates:
            try:
                return import_module(candidate)
            except ImportError as exc:
                last_error = exc
        raise ImportError(
            f"Could not import module '{module_basename}'. Tried: {', '.join(candidates)}"
        ) from last_error

    renov_module = _import_script_module("claro_renovacao_report")
    distr_module = _import_script_module("claro_distrato_report")

    ren_builder = getattr(renov_module, "build_report", None)
    dis_builder = getattr(distr_module, "generate_report", None)

    if ren_builder is None and not ren_expected.exists():
        raise AttributeError(
            "claro_renovacao_report does not expose build_report and the expected file is missing."
        )
    if dis_builder is None and not dis_expected.exists():
        raise AttributeError(
            "claro_distrato_report does not expose generate_report and the expected file is missing."
        )

    if ren_builder is None:
        ren_path = ren_expected
        ren_rows = _count_rows(ren_path)
    else:
        ren_path, ren_rows, _ = _ensure_report(ren_builder, "Renovacao", ren_expected)

    if dis_builder is None:
        dis_path = dis_expected
        dis_rows = _count_rows(dis_path)
    else:
        dis_path, dis_rows, _ = _ensure_report(dis_builder, "Distrato", dis_expected)

    out_dir = final_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    wb_final = Workbook()
    default_sheet = wb_final.active
    wb_final.remove(default_sheet)

    distrato_workbook = load_workbook(dis_path)
    for sheet_name in distrato_workbook.sheetnames:
        _copy_sheet(distrato_workbook[sheet_name], wb_final, sheet_name)

    renovacao_workbook = load_workbook(ren_path)
    for sheet_name in renovacao_workbook.sheetnames:
        _copy_sheet(renovacao_workbook[sheet_name], wb_final, sheet_name)

    wb_final.save(final_path)

    try:
        if ren_path.exists():
            ren_path.unlink()
    except Exception as exc:  # noqa: BLE001
        print(f"[WARN] Unable to remove {ren_path}: {exc}")
    try:
        if dis_path.exists():
            dis_path.unlink()
    except Exception as exc:  # noqa: BLE001
        print(f"[WARN] Unable to remove {dis_path}: {exc}")

    summary = [(final_path.name, ren_rows + dis_rows, str(final_path))]
    show_generation_popup(summary, str(out_dir))
    print(f"[OK] Consolidated file: {final_path}")
    return final_path


def run() -> Path:
    return merge_reports()


def main() -> int:
    try:
        merge_reports()
    except Exception as exc:  # noqa: BLE001
        print(f"[ERRO] {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from datetime import datetime
import re
import unicodedata
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ====== CONFIGURAÇÕES DE CAMINHOS PARA SERVIDOR ======
BASE_DIR = Path(__file__).resolve().parent.parent
MODELOS_DIR = BASE_DIR / "modelos"

REFERENCE_SHEET_NAME = "CLARO - RENOVAÇÃO 2024"
REFERENCE_HEADER_ROW = 3  # 0 = primeira linha é o header
SHEET_NAME = "Claro - Renovação"

# ====================
HEADERS_RENOVACAO: list[str] = []

SIIM_COLUMN_MAP: dict[str, list[str]] = {
    "ORDEM_SAP": ["CONTRATO", "ORDEM SAP", "ORDEM_SAP"],
    "DATA_INICIO": ["INICIO CONTRATO", "DATA INICIO", "DATA_INICIO"],
    "DATA_FIM": ["TERMINO CONTRATO", "DATA FIM", "DATA_FIM"],
    "INDICE": ["INDICE", "ÍNDICE"],
    "PROXIMO_REAJUSTE": ["DATA PROX. REAJUSTE", "PROXIMO REAJUSTE", "PROXIMO_REAJUSTE"],
    "EC": ["EC"],
    "U": ["U"],
    "D% Valor Atual": ["D% VALOR ATUAL", "D% Valor Atual", "D PERCENTUAL VALOR ATUAL"],
    "DIFERENÇA (ATUAL X FECHADO)": [
        "DIFERENCA (ATUAL X FECHADO)",
        "DIFERENÇA (ATUAL X FECHADO)",
        "DIFERENCA ATUAL X FECHADO",
    ],
    "NEGOCIADOR": ["NEGOCIADOR"],
    "STATUS": ["STATUS"],
    "PENDENCIA - NEGOCIAÇÃO": ["SITUACAO","PENDENCIA","PENDENCIA - NEGOCIACAO"],
    "OBS. NEGOCIAÇÃO": ["ULTIMO HISTORICO","OBSERVACAO","OBS. NEGOCIACAO"],
    "ALUGUEL MENSAL": ["ALUGUEL DEVIDO","ALUGUEL","ALUGUEL_MENSAL"],
    "ENDERECO": ["ENDERECO CLIENTE","ENDERECO"],
    "CENTRO_DE_CUSTO": ["CENTRO DE CUSTO","CENTRO_DE_CUSTO"],
    "CONTATO": ["TEL SOLICITANTE","CONTATO"],
    "E-MAIL": ["E-MAIL SOLICITANTE","EMAIL","E-MAIL","E_MAIL"],
    "EMPRESA_NOME": ["EMPRESA","EMPRESA NOME"],
}

SIIM_CONCLUDED_ONLY_MAP: dict[str, list[str]] = {
    "DT. STATUS NEG.": ["DATA HISTORICO", "DT STATUS NEG", "DT. STATUS NEG."],
    "INÍCIO DO NOVO ALUGUEL": ["INICIO DO NOVO ALUGUEL", "INÍCIO DO NOVO ALUGUEL"],
    "PERÍODO DA RENOVAÇÃO (Meses)": [
        "PERIODO DA RENOVACAO (MESES)",
        "PERÍODO DA RENOVAÇÃO (MESES)",
    ],
    "DATA DE INICIO DA RENOVAÇÃO": ["DATA DE INICIO DA RENOVACAO", "DATA DE INICIO DA RENOVAÇÃO"],
    "ANO ISENÇÃO DE REAJUSTE": ["ANO ISENCAO DE REAJUSTE", "ANO ISENÇÃO DE REAJUSTE"],
    "% ISENÇÃO": ["% ISENCAO", "% ISENÇÃO"],
    "NOVO INDICE NEGOCIADO": ["NOVO INDICE NEGOCIADO", "NOVO ÍNDICE NEGOCIADO"],
    "ANO ALTERAÇÃO ÍNDICE": ["ANO ALTERACAO INDICE", "ANO ALTERAÇÃO ÍNDICE"],
}

REFERENCE_PRIORITY_COLUMNS: list[str] = [
    "EMPRESA_COD", "EMPRESA_NOME", "LOCAL_NEGOCIO", "CENTRO_DE_CUSTO", "ID_GSM",
    "CLASSIFICACAO_GNI", "TP_CONTRATO", "TPC", "TIPO_DE_INFRA", "RENOVACAO_AUTOMATICA",
    "PERIODO", "ATIVO?", "NEGOCIADOR", "EMPRESA NEGOCIADORA", "CARTEIRA",
    "DT. ENVIO FORNECEDOR", "MÉDIA ARCGIS", "CONDIÇÃO CONTRATUAL", "ICG S/N",
    "MÊS ICG BASE LINE", "MÊS ICG FORECAST", "MEDIA ARCGIS 5K", "PROVISAO S/N",
    "VALOR PROVISIONADO MAIO_22", "AÇÃO JUDICIAL/TIPO", "CRITICIDADE SMART",
    "CONTRATO RE", "REGIONAL", "ENDERECO", "BAIRRO", "CIDADE", "CEP_COMPLEMENTAR",
    "ESTADO", "CONTATO", "E-MAIL", "AREA_LOCADA", "CPF", "CNPJ", "STATUS",
]

DROP_COLUMNS_BY_LETTER: tuple[str, ...] = ("BH", "BI")
STATUS_REFERENCE_LETTER = "AH"
COMBINE_COLUMN_LETTER = "AI"
COMBINE_SEPARATOR = " / "

RELNEG_TO_OUTPUT_LETTER_MAP: dict[str, str] = {
    "CW": "AO",
    "CV": "AN",
    "CX": "AR",
    "CZ": "AS",
}

NAO_CONSTA_PATTERN = re.compile(r"na[oã]\s*consta\s*no\s*siim", re.IGNORECASE)

# ====== Helpers ======
def normalize_column_key(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9]", "", text)
    return text.upper()

def find_column(df: pd.DataFrame, name: str) -> str | None:
    target = normalize_column_key(name)
    for column in df.columns:
        if normalize_column_key(column) == target:
            return column
    return None

def normalize_key_token(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits or text

def format_key_display(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()

def normalize_text_value(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()

def series_is_empty(series: pd.Series) -> pd.Series:
    mask = series.isna()
    if series.dtype == object:
        mask |= series.astype(str).str.strip() == ""
    return mask

def to_numeric_series(series: pd.Series | None) -> pd.Series | None:
    if series is None:
        return None
    if series.dtype == object:
        cleaned = series.astype(str).str.replace(r"[^\d,.\-]", "", regex=True)
        comma_mask = cleaned.str.contains(",", na=False)
        cleaned = cleaned.where(
            ~comma_mask,
            cleaned.str.replace(".", "", regex=False),
        )
        cleaned = cleaned.str.replace(",", ".", regex=False)
        return pd.to_numeric(cleaned, errors="coerce")
    return pd.to_numeric(series, errors="coerce")

def excel_index_to_letter(index: int) -> str:
    if index <= 0:
        raise ValueError("Excel column index must be positive.")
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

def build_letter_column_map(columns: list[str]) -> dict[str, str]:
    return {excel_index_to_letter(idx): name for idx, name in enumerate(columns, start=1)}

def clean_display_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    lower = text.lower()
    if not text or lower in {"nan", "<na>"}:
        return ""
    return text

def combine_series_with_separator(
    primary_series: pd.Series,
    secondary_series: pd.Series | None,
    separator: str = COMBINE_SEPARATOR,
) -> pd.Series:
    primary_values = primary_series.astype("object").map(clean_display_text)
    if secondary_series is not None:
        secondary_values = secondary_series.astype("object").map(clean_display_text)
    else:
        secondary_values = pd.Series([""] * len(primary_series), index=primary_series.index)
    combined: list[str] = []
    for primary, secondary in zip(primary_values, secondary_values):
        parts = [part for part in (primary, secondary) if part]
        combined.append(separator.join(parts))
    return pd.Series(combined, index=primary_series.index)

def copy_from_siim(
    source_df: pd.DataFrame,
    dest_df: pd.DataFrame,
    target: str,
    candidates: list[str],
    transform=None,
    series_cache: dict[str, pd.Series] | None = None,
) -> None:
    column_name = None
    for candidate in candidates:
        column_name = find_column(source_df, candidate)
        if column_name:
            break
    if not column_name:
        return
    series = source_df[column_name]
    if transform is not None:
        series = series.map(transform)
    dest_df[target] = series
    if series_cache is not None:
        series_cache[target] = series

def apply_excel_formatting_buffer(buffer: io.BytesIO, sheet_name: str) -> io.BytesIO:
    buffer.seek(0)
    wb = load_workbook(buffer)
    
    if sheet_name not in wb.sheetnames:
        return buffer
        
    ws = wb[sheet_name]
    if ws.max_row < 1 or ws.max_column < 1:
        return buffer
    
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 26

    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(fill_type="solid", fgColor="FF033E")
    navy_fill = PatternFill(fill_type="solid", fgColor="00008B")
    border_side = Side(style="thin", color="D0D0D0")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in ws[1]:
        col_idx = cell.column
        if 30 <= col_idx <= 36:
            cell.fill = navy_fill
        else:
            cell.fill = red_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal="left", vertical="top")

    sample_rows = min(ws.max_row, 200)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, sample_rows + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value = str(value)
            for part in value.splitlines():
                if len(part) > max_length:
                    max_length = len(part)
        width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    new_buffer = io.BytesIO()
    wb.save(new_buffer)
    new_buffer.seek(0)
    return new_buffer

def detect_nao_consta_rows(df: pd.DataFrame) -> pd.Series:
    mask = pd.Series(False, index=df.index)
    for col in df.columns:
        col_str = df[col].astype(str)
        mask |= col_str.str.contains(NAO_CONSTA_PATTERN)
    return mask

# ====== NÚCLEO ADAPTADO PARA API ======
def processar_relatorio_claro_renovacao(arquivo_excel_em_memoria, nseq_string: str, arquivo_modelo_em_memoria=None) -> io.BytesIO:
    """
    Função principal adaptada para receber dados da API.
    O modelo é opcional. Se não for fornecido, cria um DataFrame vazio como base.
    """
    if not nseq_string:
        raise ValueError("Nenhum NSEQ fornecido para processamento.")

    # 1. Limpa e normaliza os NSEQs recebidos
    nseq_list = [n.strip() for n in nseq_string.split(',') if n.strip()]
    rules = [normalize_key_token(n) for n in nseq_list if normalize_key_token(n)]
    rules_set = set(rules)
    order_map = {value: idx for idx, value in enumerate(rules)}

    # 2. Carrega SIIM do arquivo enviado pelo frontend
    arquivo_excel_em_memoria.seek(0)
    siim_df = pd.read_excel(arquivo_excel_em_memoria)
    siim_df.columns = [str(c).strip() for c in siim_df.columns]
    siim_letter_map = build_letter_column_map(list(siim_df.columns))

    key_col = find_column(siim_df, "NSEQ") or find_column(siim_df, "CONTRATO")
    if not key_col:
        raise KeyError("Não encontrei coluna 'NSEQ' nem 'CONTRATO' no arquivo do SIIM (RelNegociacao).")

    siim_df["__key_norm"] = siim_df[key_col].map(normalize_key_token)
    siim_df["__key_display"] = siim_df[key_col].map(format_key_display)

    # 3. LÓGICA DO MODELO OPCIONAL
    reference_df = pd.DataFrame() # Começa vazio
    
    if arquivo_modelo_em_memoria:
        arquivo_modelo_em_memoria.seek(0)
        try:
            reference_df = pd.read_excel(
                arquivo_modelo_em_memoria,
                sheet_name=REFERENCE_SHEET_NAME,
                header=REFERENCE_HEADER_ROW,
            )
            reference_df.columns = [str(c).strip() for c in reference_df.columns]
        except Exception as e:
            print(f"[WARN] Erro ao ler modelo fornecido: {e}. Prosseguindo sem modelo.")
            reference_df = pd.DataFrame()
    else:
        # Tenta achar o fixo no servidor, se não achar, segue sem modelo
        reference_file = MODELOS_DIR / "CLARO_RENOV-Report-Invest.xlsx"
        if reference_file.exists():
            try:
                reference_df = pd.read_excel(
                    reference_file,
                    sheet_name=REFERENCE_SHEET_NAME,
                    header=REFERENCE_HEADER_ROW,
                )
                reference_df.columns = [str(c).strip() for c in reference_df.columns]
            except Exception:
                pass

    # Se não conseguimos carregar nenhum modelo, criamos um DataFrame vazio com as colunas prioritárias
    if reference_df.empty:
        reference_df = pd.DataFrame(columns=REFERENCE_PRIORITY_COLUMNS + ["NSEQ - Siim"])

    # 4. Processamento Principal
    reference_columns = list(reference_df.columns)
    reference_letter_map = build_letter_column_map(reference_columns)
    combine_column_name = reference_letter_map.get(COMBINE_COLUMN_LETTER)
    status_reference_column_name = reference_letter_map.get(STATUS_REFERENCE_LETTER)

    drop_column_names: list[str] = []
    for letter in DROP_COLUMNS_BY_LETTER:
        column_name = reference_letter_map.get(letter)
        if column_name:
            drop_column_names.append(column_name)

    relneg_letter_mappings: list[tuple[str, str]] = []
    for relneg_letter, output_letter in RELNEG_TO_OUTPUT_LETTER_MAP.items():
        source_column = siim_letter_map.get(relneg_letter.upper())
        destination_column = reference_letter_map.get(output_letter.upper())
        if source_column and destination_column:
            relneg_letter_mappings.append((source_column, destination_column))

    column_am = reference_letter_map.get("AM")
    column_an = reference_letter_map.get("AN")
    column_ap = reference_letter_map.get("AP")
    column_as = reference_letter_map.get("AS")
    column_at = reference_letter_map.get("AT")
    column_au = reference_letter_map.get("AU")
    column_av = reference_letter_map.get("AV")
    column_aw = reference_letter_map.get("AW")

    mask_nao_consta = detect_nao_consta_rows(reference_df)
    ref_nao_consta = reference_df[mask_nao_consta].copy()
    ref_aguardando = reference_df[~mask_nao_consta].copy()

    ref_key_col = (
        find_column(ref_aguardando, "NSEQ - Siim")
        or find_column(ref_aguardando, "NSEQ")
        or find_column(ref_aguardando, "CONTRATO")
        or find_column(ref_aguardando, "ORDEM_SAP")
    )

    ref_keys_series = pd.Series(dtype="object", index=ref_aguardando.index)
    ref_keys_unique: list[str] = []
    if ref_key_col is not None:
        ref_keys_series = ref_aguardando[ref_key_col].map(normalize_key_token).fillna("")
        seen_ref_keys: set[str] = set()
        for key in ref_keys_series:
            if not key or key in seen_ref_keys:
                continue
            seen_ref_keys.add(key)
            ref_keys_unique.append(key)

    status_reference_map: pd.Series | None = None
    if status_reference_column_name and status_reference_column_name in ref_aguardando.columns and not ref_aguardando.empty:
        ref_status = ref_aguardando[[status_reference_column_name]].copy()
        ref_status["__key_norm"] = ref_keys_series.reindex(ref_status.index).fillna("")
        ref_status = ref_status[ref_status["__key_norm"].astype(bool)]
        if not ref_status.empty:
            status_reference_map = ref_status.set_index("__key_norm")[status_reference_column_name]

    status_reference_lookup = status_reference_map.to_dict() if status_reference_map is not None else None

    next_order = len(order_map)
    seen_keys = set(order_map)
    for key in ref_keys_unique:
        if key not in seen_keys:
            order_map[key] = next_order
            seen_keys.add(key)
            next_order += 1

    target_keys = rules_set
    siim_filtered = siim_df[siim_df["__key_norm"].isin(target_keys)].copy()

    if siim_filtered.empty:
        raise ValueError("Nenhuma linha do SIIM coincide com os NSEQs informados.")

    siim_filtered["__order"] = siim_filtered["__key_norm"].map(order_map)
    siim_filtered = siim_filtered.sort_values("__order").reset_index(drop=True)

    rep_aguardando = pd.DataFrame()
    rep_aguardando_keys: pd.Series | None = None
    siim_series_cache: dict[str, pd.Series] = {}
    siim_part_for_output: pd.DataFrame | None = None

    if not siim_filtered.empty:
        siim_part = siim_filtered.copy()
        rep_aguardando = pd.DataFrame(index=siim_part.index)
        siim_part_for_output = siim_part
        rep_aguardando_keys = siim_part["__key_norm"].copy()

        for dest_col, candidates in SIIM_COLUMN_MAP.items():
            copy_from_siim(siim_part, rep_aguardando, dest_col, candidates, series_cache=siim_series_cache)

        situacao_col = find_column(siim_part, "SITUACAO") or find_column(siim_part, "STATUS")
        concluido_mask: pd.Series | None = None
        if situacao_col is not None:
            situacao_series = siim_part[situacao_col].map(normalize_text_value)
            concluido_mask = situacao_series.str.contains("concluid", na=False)

        for dest_col, candidates in SIIM_CONCLUDED_ONLY_MAP.items():
            copy_from_siim(siim_part, rep_aguardando, dest_col, candidates, series_cache=siim_series_cache)
            if dest_col not in rep_aguardando.columns:
                continue
            if concluido_mask is not None:
                rep_aguardando[dest_col] = rep_aguardando[dest_col].where(concluido_mask, pd.NA)
            else:
                rep_aguardando[dest_col] = pd.NA

        if not ref_aguardando.empty:
            ref_aug = ref_aguardando.copy()
            ref_aug["__key_norm"] = ref_keys_series.reindex(ref_aug.index).fillna("")
            for col in ref_aug.columns:
                if col == "__key_norm":
                    continue
                if col not in rep_aguardando.columns:
                    rep_aguardando[col] = pd.NA

            for col in REFERENCE_PRIORITY_COLUMNS:
                if col not in ref_aug.columns:
                    continue
                series_map = ref_aug.set_index("__key_norm")[col]
                mapped = siim_part["__key_norm"].map(series_map)
                if col not in rep_aguardando.columns:
                    rep_aguardando[col] = mapped
                else:
                    base = rep_aguardando[col].astype("object") if col in rep_aguardando.columns else pd.Series(pd.NA, index=rep_aguardando.index)
                    mapped = mapped.where(~series_is_empty(mapped), pd.NA).astype("object")
                    rep_aguardando[col] = mapped.combine_first(base)

        rep_aguardando["NSEQ - Siim"] = siim_part["__key_display"]

        for source_column, destination_column in relneg_letter_mappings:
            if source_column in siim_part.columns:
                rep_aguardando[destination_column] = siim_part[source_column].reindex(rep_aguardando.index)

    rep_nao_consta = ref_nao_consta.copy()

    if HEADERS_RENOVACAO:
        final_cols = HEADERS_RENOVACAO[:]
    else:
        ref_cols = list(reference_df.columns)
        siim_extra = [c for c in rep_aguardando.columns if c not in ref_cols]
        final_cols = ref_cols + siim_extra

    columns_to_remove = [name for name in drop_column_names if name and name in final_cols]
    if columns_to_remove:
        final_cols = [col for col in final_cols if col not in columns_to_remove]
        rep_nao_consta = rep_nao_consta.drop(columns=columns_to_remove, errors='ignore')
        if not rep_aguardando.empty:
            rep_aguardando = rep_aguardando.drop(columns=columns_to_remove, errors='ignore')

    def ensure_columns(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
        for c in cols:
            if c not in df.columns:
                df[c] = pd.NA
        return df[cols]

    rep_nao_consta = ensure_columns(rep_nao_consta, final_cols)
    rep_aguardando = ensure_columns(rep_aguardando, final_cols) if not rep_aguardando.empty else pd.DataFrame(columns=final_cols)

    # 
    # CÁLCULOS COMPLEXOS DE COLUNAS (AQUI ONDE CORTOU)
    # 
    if not rep_aguardando.empty and siim_part_for_output is not None:
        index = rep_aguardando.index

        def reindexed_siim_series(letter: str) -> pd.Series | None:
            column = siim_letter_map.get(letter.upper())
            if not column or column not in siim_part_for_output.columns:
                return None
            return siim_part_for_output[column].reindex(index)

        def numeric_series_from_letter(letter: str) -> pd.Series | None:
            return to_numeric_series(reindexed_siim_series(letter))

        def datetime_series_from_letter(letter: str) -> pd.Series | None:
            series = reindexed_siim_series(letter)
            if series is None:
                return None
            return pd.to_datetime(series, errors="coerce")

        aluguel_series = numeric_series_from_letter("U")
        cv_series = numeric_series_from_letter("CV")
        bw_series = reindexed_siim_series("BW")
        by_series = reindexed_siim_series("BY")
        ee_series = reindexed_siim_series("EE")
        dr_series = reindexed_siim_series("DR")
        cj_series = reindexed_siim_series("CJ")

        bw_mask = (
            bw_series.astype(str).str.strip().str.lower().isin({"sim", "yes"})
            if bw_series is not None
            else None
        )

        if column_am and column_am in rep_aguardando.columns and aluguel_series is not None:
            mask = aluguel_series.notna()
            rep_aguardando.loc[mask, column_am] = aluguel_series[mask]

        if column_an and column_an in rep_aguardando.columns and aluguel_series is not None and cv_series is not None:
            novo_valor_series = aluguel_series - cv_series
            mask = novo_valor_series.notna()
            rep_aguardando.loc[mask, column_an] = novo_valor_series[mask]

        if column_ap and column_ap in rep_aguardando.columns:
            am_numeric = to_numeric_series(rep_aguardando[column_am]) if column_am else None
            an_numeric = to_numeric_series(rep_aguardando[column_an]) if column_an else None
            if am_numeric is not None and an_numeric is not None:
                diferenca_ap = am_numeric - an_numeric
                mask = diferenca_ap.notna()
                rep_aguardando.loc[mask, column_ap] = diferenca_ap[mask]

        if column_at and column_at in rep_aguardando.columns:
            t_series = datetime_series_from_letter("T")
            if t_series is not None:
                next_day = t_series + pd.to_timedelta(1, unit="D")
                mask = next_day.notna()
                rep_aguardando.loc[mask, column_at] = next_day[mask]

        if column_at and column_at in rep_aguardando.columns and bw_mask is not None and ee_series is not None:
            override_values = ee_series.reindex(index)
            mask = bw_mask.reindex(index, fill_value=False) & override_values.notna()
            rep_aguardando.loc[mask, column_at] = override_values[mask]

        if column_au and column_au in rep_aguardando.columns:
            cm_series = reindexed_siim_series("CM")
            month_series = datetime_series_from_letter("CX")
            year_series = datetime_series_from_letter("V")

            if cm_series is not None and month_series is not None and year_series is not None:
                cm_mask = cm_series.astype(str).str.strip().str.lower().isin({"sim", "yes"})
                updated_au = rep_aguardando[column_au].astype("object").copy()

                for idx in index:
                    if idx not in cm_mask.index or not bool(cm_mask.loc[idx]):
                        continue
                    month_value = month_series.loc[idx]
                    year_value = year_series.loc[idx]

                    if pd.isna(month_value) or pd.isna(year_value):
                        continue

                    month_number = int(pd.Timestamp(month_value).month)
                    year_number = int(pd.Timestamp(year_value).year)
                    formatted = f"{month_number:02d}/{year_number}"
                    updated_au.at[idx] = formatted

                rep_aguardando[column_au] = updated_au

                if column_av and column_av in rep_aguardando.columns:
                    non_empty_mask = (~updated_au.isna()) & (updated_au.astype(str).str.strip() != "")
                    rep_aguardando.loc[non_empty_mask, column_av] = "100%"
                    rep_aguardando.loc[~non_empty_mask, column_av] = pd.NA

        if column_as and column_as in rep_aguardando.columns and bw_mask is not None and by_series is not None:
            source_values = by_series.reindex(index)
            mask = bw_mask.reindex(index, fill_value=False) & source_values.notna()
            rep_aguardando.loc[mask, column_as] = source_values[mask]

        if column_aw and column_aw in rep_aguardando.columns and dr_series is not None and cj_series is not None:
            dr_mask = dr_series.astype(str).str.strip().str.lower().str.contains("concluid", na=False)
            override_values = cj_series.reindex(index)
            mask = dr_mask.reindex(index, fill_value=False) & override_values.notna()
            rep_aguardando.loc[mask, column_aw] = override_values[mask]

    if combine_column_name and not rep_aguardando.empty and combine_column_name in rep_aguardando.columns:
        siim_series = siim_series_cache.get(combine_column_name)
        if siim_series is not None:
            siim_series = siim_series.reindex(rep_aguardando.index)
            status_series = None
            if status_reference_lookup and rep_aguardando_keys is not None:
                status_series = rep_aguardando_keys.map(status_reference_lookup).reindex(rep_aguardando.index)
            rep_aguardando[combine_column_name] = combine_series_with_separator(siim_series, status_series)

    # 5. Finalização e Concatenação
    rep_nao_consta["__ORIGEM__"] = "MODELO (não consta no SIIM)"
    if not rep_aguardando.empty:
        rep_aguardando["__ORIGEM__"] = "SIIM (RelNegociacao)"

    result = pd.concat([rep_aguardando, rep_nao_consta], ignore_index=True)
    result = result.drop(columns=["__ORIGEM__"], errors="ignore")

    # 6. Salvar e Formatar em Memória
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    output_formatado = apply_excel_formatting_buffer(output, SHEET_NAME)

    return output_formatado

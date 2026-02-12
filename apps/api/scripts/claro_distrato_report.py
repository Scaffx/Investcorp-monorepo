#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from pathlib import Path
from datetime import datetime, timedelta
import sys
import re
import unicodedata
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

try:
    from .utils import show_generation_popup
except ImportError:
    current_dir = Path(__file__).resolve().parent
    if str(current_dir) not in sys.path:
        sys.path.append(str(current_dir))
    from utils import show_generation_popup  # type: ignore[import-not-found]
# ====== CONFIG ======
DESKTOP = Path.home() / "Desktop"
BASE_DIR = DESKTOP / "Report"
MODELOS_DIR = BASE_DIR / "Modelos"
INPUT_FILE = MODELOS_DIR / "RelNegociacao.xlsx"
REGRAS_DIR = BASE_DIR / "REGRAS"
RULES_FILE = REGRAS_DIR / "Claro_Distrato_regras.txt"
REFERENCE_FILE = MODELOS_DIR / "CLARO_DISTRATO-Report-Invest.xlsx"
REFERENCE_SHEET_NAME = "CLARO - DISTRATO"
REFERENCE_HEADER_ROW = 3
SHEET_NAME = "Claro - Distrato"
# ====================
# Cabeçalho final do report (ORDEM FINAL) — a ordem aqui é a que você quer ver no Excel.
HEADERS_DISTRATO = [
    "EMPRESA_COD","EMPRESA_NOME","ORDEM_SAP","CONTRATO RE","LOCAL_NEGOCIO","CENTRO_DE_CUSTO",
    "REGIONAL","ID_GSM","CLASSIFICACAO_GNI","TP_CONTRATO","TPC","TIPO_DE_INFRA",
    "DATA_INICIO","DATA_FIM","RENOVACAO_AUTOMATICA","PERIODO","INDICE","PROXIMO_REAJUSTE",
    "ATIVO?","NOME_LOCADOR","ENDERECO","BAIRRO","CIDADE","CEP_COMPLEMENTAR","ESTADO",
    "CONTATO","E-MAIL","AREA_LOCADA","CPF","CNPJ","NEGOCIADOR","EMPRESA NEGOCIADORA",
    "CARTEIRA","STATUS","PENDENCIA - NEGOCIAÇÃO","DT. ENVIO FORNECEDOR",
    "DT. STATUS NEG.","OBS. NEGOCIAÇÃO","ALUGUEL MENSAL"
]
# Do SIIM (RelNegociacao) no DISTRATO (pelas suas regras amarelas)
SIIM_COLS_DISTRATO = [
    "ORDEM_SAP","DATA_INICIO","DATA_FIM","INDICE","PROXIMO_REAJUSTE",
    "NOME_LOCADOR","E-MAIL","STATUS","PENDENCIA - NEGOCIAÇÃO","DT. STATUS NEG.",
    "OBS. NEGOCIAÇÃO","ALUGUEL MENSAL"
]
REFERENCE_PRIORITY_COLUMNS = [
    "EMPRESA_COD","EMPRESA_NOME","LOCAL_NEGOCIO","CENTRO_DE_CUSTO","REGIONAL","ID_GSM",
    "CLASSIFICACAO_GNI","TP_CONTRATO","TPC","TIPO_DE_INFRA","RENOVACAO_AUTOMATICA",
    "PERIODO","ATIVO?","NEGOCIADOR","EMPRESA NEGOCIADORA","CARTEIRA","DT. ENVIO FORNECEDOR"
]
SIIM_COLUMN_MAP = {
    "ORDEM_SAP": ["CONTRATO"],
    "DATA_INICIO": ["INICIO CONTRATO"],
    "DATA_FIM": ["TERMINO CONTRATO"],
    "INDICE": ["INDICE"],
    "PROXIMO_REAJUSTE": ["DATA PROX. REAJUSTE"],
    "NEGOCIADOR": ["NEGOCIADOR"],
    "STATUS": ["STATUS"],
    "PENDENCIA - NEGOCIAÇÃO": ["SITUACAO"],
    "DT. STATUS NEG.": ["DATA HISTORICO"],
    "OBS. NEGOCIAÇÃO": ["ULTIMO HISTORICO"],
    "ALUGUEL MENSAL": ["ALUGUEL DEVIDO"],
    "ENDERECO": ["ENDERECO CLIENTE"],
    "CENTRO_DE_CUSTO": ["CENTRO DE CUSTO"],
    "CONTATO": ["TEL SOLICITANTE"],
    "E-MAIL": ["E-MAIL SOLICITANTE"],
    "EMPRESA_NOME": ["EMPRESA"]
}
def normalize_column_key(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9]", "", text)
    return text.upper()
def find_column(df: pd.DataFrame, name: str) -> str | None:
    target = normalize_column_key(name)
    for column in df.columns:
        if normalize_column_key(column) == target:
            return column
    return None
def normalize_nseq(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            value = int(value)
    text = str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits or text
def format_nseq_display(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()
def load_nseq_rules(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo de regras não encontrado: {path}")

    ordered: list[str] = []
    seen: set[str] = set()
    line_index = 0

    with path.open("r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue

            parts = re.split(r"[;,\s]+", line)
            tokens: list[str] = []
            for part in parts:
                norm = normalize_nseq(part)
                if not norm or not norm.isdigit():
                    continue
                tokens.append(norm)

            if not tokens:
                continue

            line_index += 1
            if len(tokens) > 1 and tokens[0] == str(line_index):
                tokens = tokens[1:]

            for norm in tokens:
                if norm in seen:
                    continue
                ordered.append(norm)
                seen.add(norm)

    if not ordered:
        raise ValueError(f"Nenhum NSEQ válido encontrado no arquivo de regras: {path}")
    return ordered
def series_is_empty(series: pd.Series) -> pd.Series:
    mask = series.isna()
    if series.dtype == object:
        mask |= series.astype(str).str.strip() == ""
    return mask
def copy_from_siim(source_df: pd.DataFrame, dest_df: pd.DataFrame, target: str, candidates: list[str], transform=None) -> None:
    column_name = None
    for candidate in candidates:
        column_name = find_column(source_df, candidate)
        if column_name:
            break
    if not column_name:
        return
    series = source_df[column_name]
    if transform is not None:
        series = series.map(transform)
    dest_df[target] = series


def apply_excel_formatting(path: Path, sheet_name: str) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb[sheet_name]
    if ws.max_row < 1 or ws.max_column < 1:
        wb.save(path)
        return

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 26

    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(fill_type="solid", fgColor="FF033E")
    navy_fill = PatternFill(fill_type="solid", fgColor="00008B")
    border_side = Side(style="thin", color="D0D0D0")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in ws[1]:
        col_idx = cell.column
        if 34 <= col_idx <= 38:
            cell.fill = navy_fill
        elif col_idx == 39:
            cell.fill = red_fill
        else:
            cell.fill = red_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal="left", vertical="top")

    sample_rows = min(ws.max_row, 200)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, sample_rows + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value = str(value)
            for part in value.splitlines():
                if len(part) > max_length:
                    max_length = len(part)
        width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(path)
def generate_report() -> tuple[Path, int, Path]:
    yesterday = datetime.now() - timedelta(days=1)
    folder_token = yesterday.strftime("%d%m%Y")
    out_dir = BASE_DIR / folder_token
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"CLARO_DISTRATO_{folder_token}.xlsx"
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Arquivo de entrada não encontrado: {INPUT_FILE}")
    if INPUT_FILE.suffix.lower() in (".xlsx", ".xls"):
        siim_df = pd.read_excel(INPUT_FILE)
    else:
        siim_df = pd.read_csv(INPUT_FILE, sep=";", encoding="utf-8", engine="python")
        if siim_df.shape[1] == 1:
            siim_df = pd.read_csv(INPUT_FILE, sep=",", encoding="utf-8", engine="python")
    siim_df.columns = [str(c).strip() for c in siim_df.columns]
    rules = load_nseq_rules(RULES_FILE)
    nseq_column = find_column(siim_df, "NSEQ")
    if not nseq_column:
        raise KeyError("Coluna 'NSEQ' não encontrada em RelNegociacao.xlsx")
    siim_df["__nseq_norm"] = siim_df[nseq_column].map(normalize_nseq)
    siim_df["NSEQ - Siim"] = siim_df["__nseq_norm"].map(format_nseq_display)
    order_map = {value: idx for idx, value in enumerate(rules)}
    filtered = siim_df[siim_df["__nseq_norm"].isin(order_map)].copy()
    if filtered.empty:
        raise ValueError(
            "Nenhuma linha do RelNegociacao.xlsx corresponde aos NSEQ definidos em Claro_Distrato_regras.txt."
        )
    filtered["__order"] = filtered["__nseq_norm"].map(order_map)
    filtered = filtered.sort_values("__order").reset_index(drop=True)
    matched = set(filtered["__nseq_norm"])
    missing = [value for value in rules if value not in matched]
    if missing:
        print(
            f"[WARN] NSEQ não encontrados no RelNegociacao.xlsx: {', '.join(missing)}",
            file=sys.stderr,
        )
    filtered["NSEQ - Siim"] = filtered["__nseq_norm"].map(format_nseq_display)
    rep = pd.DataFrame(index=filtered.index)
    rep["__nseq_norm"] = filtered["__nseq_norm"]
    for dest_col, candidates in SIIM_COLUMN_MAP.items():
        copy_from_siim(filtered, rep, dest_col, candidates)
    rep["NSEQ - Siim"] = filtered["NSEQ - Siim"]
    if not REFERENCE_FILE.exists():
        raise FileNotFoundError(f"Arquivo de referência não encontrado: {REFERENCE_FILE}")
    reference_df = pd.read_excel(
        REFERENCE_FILE,
        sheet_name=REFERENCE_SHEET_NAME,
        header=REFERENCE_HEADER_ROW,
    )
    reference_df.columns = [str(c).strip() for c in reference_df.columns]
    if "NSEQ - Siim" not in reference_df.columns:
        raise KeyError("Coluna 'NSEQ - Siim' não encontrada na planilha de referência.")
    reference_df = reference_df.dropna(subset=["NSEQ - Siim"]).copy()
    reference_df["__nseq_norm"] = reference_df["NSEQ - Siim"].map(normalize_nseq)
    reference_df = reference_df[reference_df["__nseq_norm"] != ""].copy()
    reference_df = reference_df.drop_duplicates("__nseq_norm", keep="first").set_index("__nseq_norm")
    priority_columns = set(REFERENCE_PRIORITY_COLUMNS)
    for column in reference_df.columns:
        if column == "__nseq_norm" or column not in HEADERS_DISTRATO:
            continue
        series = reference_df[column]
        mapped = rep["__nseq_norm"].map(series)
        mapped = mapped.where(~series_is_empty(mapped), pd.NA).astype("object")
        if column in priority_columns:
            base = rep[column] if column in rep.columns else pd.Series(pd.NA, index=rep.index)
            base = base.astype("object")
            rep[column] = mapped.combine_first(base)
        else:
            if column in rep.columns:
                rep[column] = rep[column].astype("object")
                mask = series_is_empty(rep[column])
                rep.loc[mask, column] = mapped[mask]
            else:
                rep[column] = mapped
    rep = rep.drop(columns=["__nseq_norm"], errors="ignore")
    rep = rep.drop(columns=["NSEQ - Siim"], errors="ignore")
    for column in HEADERS_DISTRATO:
        if column not in rep.columns:
            rep[column] = pd.NA
    rep = rep[HEADERS_DISTRATO]
    row_count = len(rep.index)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        rep.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    apply_excel_formatting(out_path, SHEET_NAME)

    return out_path, row_count, out_dir
def main() -> int:
    try:
        out_path, row_count, out_dir = generate_report()
    except Exception as exc:  # noqa: BLE001 - expor mensagem ao usuário
        print(f"[ERRO] {exc}", file=sys.stderr)
        return 1
    print(f"[OK] Distrato gerado: {out_path}")
    summary = [(out_path.name, row_count, str(out_path))]
    show_generation_popup(summary, str(out_dir))
    return 0

if __name__ == "__main__":
    sys.exit(main())

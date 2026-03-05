#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Casas Bahia Report - Adaptado para API Django/Railway
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =========================
# CONFIGURAÇÕES DE PASTA (SERVIDOR)
# =========================
BASE_DIR = Path(__file__).resolve().parent.parent
REPORT_DIR = BASE_DIR / "reports_gerados"
MODELOS_DIR = BASE_DIR / "modelos"

DEFAULT_SHEET_NAME = "BASE_CASAS_BAHIA"
OUTPUT_SHEET_NAME = "Report_Casas_Bahia"

YELLOW_COLUMNS = [
    "FOLLOW UP",
    "OBS P/ CONSULTORIA",
    "CONTATO TELEFONE",
    "CONTATO E-MAIL",
    "MOTIVO RECUSADO",
    "Coluna1",
    "NSEQ",
]

RELNEG_SOURCE_CANDIDATES: Dict[str, List[str]] = {
    "CONTRATO": ["Junção", "Juncao"],
    "CONTRATO ANT.": ["CENTRO DE CUSTO"],
    "CED": ["CONTRATO"],
    "# CT": ["ONDA"],
    "BAND.": ["BANDEIRA"],
    "DENOMINACAO": ["DENOMINACAO/ NOME"],
    "ENDERECO": ["ENDEREÇO CLIENTE", "ENDERECO CLIENTE"],
    "CIDADE": ["Cidade"],
    "UF": ["UF"],
    "TIPO": ["TIPO IMOVEL", "TIPO DO LOGRADOURO", "TIPO"],
    "AREA TOTAL (M2)": ["M  AREA TOTAL", "M2 AREA TOTAL"],
    "AREA DE VENDA (M2)": ["M2 área Venda", "M2 AREA VENDA"],
    "PLANO GI": ["Obs Premissa"],
    "NEGOCIADOR": ["Negociador"],
    "NEGOCIADOR GERAL": ["Resp. Adm."],
    "ALUGUEL LOJA": ["ALUGUEL DEVIDO", "Valor Negociado"],
    "ALUGUEL CT": ["ALUGUEL DEVIDO"],
    "ALU %": ["% Aluguel Variavel Atual"],
    "INDICE": ["INDICE"],
    "DT BASE": ["DATA PRÓX. REAJUSTE", "DATA PROX. REAJUSTE"],
    "DT FIM CT": ["TERMINO CONTRATO"],
    "INICIO NEGOCIACAO": ["Data Início Negociação", "DATA INICIO NEGOCIACAO"],
    "ULTIMA ABORDAGEM": ["Data Historico", "DATA HISTORICO"],
    "FIM NEGOCIACAO": ["Data Conclusao", "DATA CONCLUSAO", "Data Conclusão"],
    "STATUS": ["Status", "Situação"],
    "PROPOSTA APRESENTADA": ["Proposta"],
    "MOTIVO RECUSADO": ["Motivadores sem Exito", "MOTIVADORES SEM EXITO"],
    "NEGOCIACAO FINALIZADA": ["Contra Proposta", "CONTRA PROPOSTA"],
    "FOLLOW UP": ["Ultimo Historico", "ULTIMO HISTORICO"],
    "OBS P/ CONSULTORIA": ["Observação", "OBSERVACAO"],
    "CONTATO TELEFONE": ["Tel Solicitante"],
    "CONTATO E-MAIL": ["E-mail Solicitante"],
    "NSEQ": ["NSEQ"],
}

DATE_COLUMNS = [
    "DT BASE",
    "DT FIM CT",
    "INICIO NEGOCIACAO",
    "ULTIMA ABORDAGEM",
    "FIM NEGOCIACAO",
]

DATE_COLUMNS_NORM = {normalize_value(name) for name in DATE_COLUMNS}
RELNEG_ONLY_COLUMNS = {"MOTIVO RECUSADO", "NEGOCIAÇÃO FINALIZADA", "FIM NEGOCIAÇÃO"}
RELNEG_ONLY_COLUMNS_NORM = {normalize_value(name) for name in RELNEG_ONLY_COLUMNS}

def normalize_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
    else:
        if pd.isna(value):
            return ""
        if isinstance(value, float):
            if value.is_integer():
                text = str(int(value))
            else:
                text = format(value, "g")
        elif isinstance(value, int):
            text = str(value)
        else:
            text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.upper()

def normalize_nseq_value(value: object) -> str:
    text = normalize_value(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits or text

def is_final_status_value(value: object) -> bool:
    norm = normalize_value(value)
    if not norm:
        return False
    return norm in {"FINALIZADO", "FINALIZADA", "RECUSADO", "RECUSADO VIA"}

def is_blank_value(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        return stripped == "" or stripped.lower() in {"nan", "nat"}
    return pd.isna(value)

def series_is_blank(series: pd.Series) -> pd.Series:
    return series.apply(is_blank_value)

def series_is_filled(series: pd.Series) -> pd.Series:
    return ~series_is_blank(series)

def format_date_string(value: object) -> str:
    if is_blank_value(value):
        return ""
    try:
        parsed = pd.to_datetime(value, dayfirst=True, errors="raise")
        if pd.isna(parsed):
            return str(value)
        return parsed.strftime("%d/%m/%Y")
    except Exception:
        try:
            parsed = pd.to_datetime(value, dayfirst=False, errors="raise")
            if pd.isna(parsed):
                return str(value)
            return parsed.strftime("%d/%m/%Y")
        except Exception:
            return str(value)

COLUMN_ALIASES: Dict[str, List[str]] = {
    "NSEQ": ["NSEQ", "NSEQ_ESCOLHIDO"],
    "CONTRATO": ["CONTRATO"],
    "JUNÇÃO": ["JUNÇÃO", "JUNCAO", "CHAVE"],
    "CENTRO DE CUSTO": ["CENTRO DE CUSTO", "CENTRODECUSTO", "CC"],
}

COLUMN_ALIAS_MAP: Dict[str, str] = {}
for canonical, aliases in COLUMN_ALIASES.items():
    canonical_norm = normalize_value(canonical)
    COLUMN_ALIAS_MAP[canonical_norm] = canonical_norm
    for alias in aliases:
        COLUMN_ALIAS_MAP[normalize_value(alias)] = canonical_norm

DEFAULT_RULE_COLUMNS = ["NSEQ", "Junção", "CONTRATO", "CENTRO DE CUSTO"]
DEFAULT_RULE_COLUMNS_NORM = [normalize_value(col) for col in DEFAULT_RULE_COLUMNS]

@dataclass(frozen=True)
class Rule:
    raw: str
    normalized: str
    columns: Optional[Sequence[str]] = None

def load_spec(spec_path: Path, sheet_name: str = DEFAULT_SHEET_NAME):
    xl = pd.ExcelFile(spec_path)
    if sheet_name not in xl.sheet_names:
        raise ValueError(f'Aba "{sheet_name}" não encontrada em {spec_path.name}.')
    spec_df = pd.read_excel(spec_path, sheet_name=sheet_name)
    header_map = spec_df.iloc[0].to_dict()
    final_order = list(header_map.values())
    ref_df = spec_df.iloc[1:].reset_index(drop=True)
    return header_map, final_order, ref_df

def load_reference_report(spec_path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(spec_path, header=1)
    except Exception as exc:
        raise RuntimeError(f"Não foi possível ler o report base '{spec_path.name}': {exc}") from exc

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def apply_status_depara(series: pd.Series) -> pd.Series:
    depara_norm: Dict[str, str] = {
        normalize_value("A Negociar"): "À NEGOCIAR",
        normalize_value("Em Negociação"): "EM NEGOCIAÇÃO",
        normalize_value("Aguardando Minuta"): "EM NEGOCIAÇÃO",
        normalize_value("Em Assinatura"): "EM NEGOCIAÇÃO",
        normalize_value("Finalizado Sem Sucesso"): "RECUSADO",
        normalize_value("Concluído"): "FINALIZADO",
        normalize_value("Cancelado"): "RECUSADO VIA",
    }
    def _map(val: object) -> object:
        norm = normalize_value(val)
        return depara_norm.get(norm, val)
    return series.map(_map)

def build_report(
    header_map: Dict[str, str],
    final_order: Sequence[str],
    relneg_df: pd.DataFrame,
    base_report_df: pd.DataFrame,
    rules: Sequence[Rule],
    ref_df: pd.DataFrame,
) -> pd.DataFrame:
    forward_map: Dict[str, str] = {}
    reverse_map: Dict[str, str] = {}
    for src, tgt in header_map.items():
        if not isinstance(src, str) or not isinstance(tgt, str):
            continue
        forward_map[src] = tgt
        reverse_map[tgt] = src

    rename_map = {src: tgt for src, tgt in forward_map.items() if src in relneg_df.columns}
    normalized_relneg_columns = {normalize_value(col): col for col in relneg_df.columns}
    relneg_named = relneg_df.rename(columns=rename_map).copy()

    for final_col in final_order:
        norm_final = normalize_value(final_col)
        candidates = RELNEG_SOURCE_CANDIDATES.get(norm_final)
        if not candidates:
            continue
        for candidate in candidates:
            source_name = normalized_relneg_columns.get(normalize_value(candidate))
            if not source_name:
                continue
            relneg_named[final_col] = relneg_df[source_name]
            break

    if "STATUS" in relneg_named.columns:
        relneg_named["STATUS"] = apply_status_depara(relneg_named["STATUS"])

    for col in final_order:
        if col not in relneg_named.columns:
            relneg_named[col] = pd.NA

    relneg_named["NSEQ"] = relneg_named.get("NSEQ")
    relneg_named["__key"] = relneg_named["NSEQ"].map(normalize_nseq_value)
    relneg_named = relneg_named.dropna(subset=["__key"]).drop_duplicates("__key", keep="first").set_index("__key")

    base_data = base_report_df.copy()
    for col in final_order:
        if col not in base_data.columns:
            base_data[col] = pd.NA
        elif normalize_value(col) in RELNEG_ONLY_COLUMNS_NORM:
            base_data[col] = pd.NA

    if "NSEQ" not in base_data.columns:
        base_data["NSEQ"] = pd.NA

    base_data = base_data[final_order]
    base_data["__key"] = base_data["NSEQ"].map(normalize_nseq_value)
    base_indexed = base_data.dropna(subset=["__key"]).drop_duplicates("__key", keep="first").set_index("__key")

    rule_keys: List[str] = []
    for rule in rules:
        key = normalize_nseq_value(rule.normalized or rule.raw)
        if key and key not in rule_keys:
            rule_keys.append(key)

    final_df = base_indexed.reindex(rule_keys)
    if final_df.empty:
        final_df = pd.DataFrame(index=rule_keys, columns=final_order, dtype=object)
    else:
        final_df = final_df[final_order].copy()
    final_df = final_df.astype(object)

    relneg_indexed = relneg_named.reindex(rule_keys)
    for col in final_order:
        if col not in relneg_indexed.columns:
            continue
        source_col = relneg_indexed[col]
        filled_mask = series_is_filled(source_col)
        col_norm = normalize_value(col)

        if col_norm == normalize_value("NEGOCIACAO FINALIZADA"):
            status_series = relneg_indexed.get("STATUS")
            status_mask = status_series.map(is_final_status_value) if status_series is not None else filled_mask
            mask = status_mask & filled_mask
            if mask.any():
                final_df.loc[mask, col] = source_col[mask]
            continue

        if col in YELLOW_COLUMNS:
            if filled_mask.any():
                final_df.loc[filled_mask, col] = source_col[filled_mask]
        elif col_norm in DATE_COLUMNS_NORM:
            if filled_mask.any():
                final_df.loc[filled_mask, col] = source_col[filled_mask]
        else:
            blank_mask = series_is_blank(final_df[col])
            mask = blank_mask & filled_mask
            if mask.any():
                final_df.loc[mask, col] = source_col[mask]

    if not ref_df.empty:
        ref_named = ref_df.rename(columns=rename_map)
        for col in final_order:
            if normalize_value(col) in RELNEG_ONLY_COLUMNS_NORM:
                continue
            if col not in ref_named.columns:
                continue
            defaults = ref_named[col].dropna()
            if defaults.empty:
                continue
            default_value = defaults.iloc[0]
            mask = series_is_blank(final_df[col])
            if mask.any() and not is_blank_value(default_value):
                final_df.loc[mask, col] = default_value

    final_df = final_df.reindex(columns=final_order)
    final_df.reset_index(drop=True, inplace=True)

    neg_final_col = next((col for col in final_df.columns if normalize_value(col) == "NEGOCIACAO FINALIZADA"), None)
    if neg_final_col and "STATUS" in final_df.columns:
        status_mask = final_df["STATUS"].map(is_final_status_value)
        final_df.loc[~status_mask, neg_final_col] = pd.NA

    for col in final_df.columns:
        if normalize_value(col) in DATE_COLUMNS_NORM:
            final_df[col] = final_df[col].apply(format_date_string)

    return final_df

def filter_by_rules(df: pd.DataFrame, rules: Sequence[Rule]) -> pd.DataFrame:
    if df.empty:
        raise ValueError("A base de negociação está vazia.")
    if not rules:
        raise ValueError("Nenhuma regra fornecida para filtragem.")

    columns_by_norm: Dict[str, List[str]] = {}
    for col in df.columns:
        columns_by_norm.setdefault(normalize_value(col), []).append(col)

    normalized_cache: Dict[str, pd.Series] = {}
    def get_series(col_name: str) -> pd.Series:
        if col_name not in normalized_cache:
            normalized_cache[col_name] = df[col_name].fillna("").astype(str).map(normalize_value)
        return normalized_cache[col_name]

    order_map: Dict[str, int] = {}
    for idx, rule in enumerate(rules):
        if rule.normalized and rule.normalized not in order_map:
            order_map[rule.normalized] = idx

    selected: List[tuple[int, str]] = []
    used_indices: Set[int] = set()
    matched_rules: Set[str] = set()
    missing: List[str] = []

    for rule in rules:
        norm_value = rule.normalized
        if not norm_value or norm_value in matched_rules:
            continue

        desired_norms = list(rule.columns or [])
        desired_norms.extend(DEFAULT_RULE_COLUMNS_NORM)
        candidate_columns: List[str] = []
        seen_norms: Set[str] = set()

        for col_norm in desired_norms:
            if not col_norm or col_norm in seen_norms:
                continue
            seen_norms.add(col_norm)
            candidate_columns.extend(columns_by_norm.get(col_norm, []))

        if not candidate_columns:
            missing.append(f"{rule.raw} (colunas indisponíveis)")
            continue

        found_index: Optional[int] = None
        for column_name in candidate_columns:
            series = get_series(column_name)
            match_indices = series[series == norm_value].index
            for idx in match_indices:
                if idx not in used_indices:
                    found_index = idx
                    break
            if found_index is not None:
                break

        if found_index is None:
            missing.append(rule.raw)
            continue

        used_indices.add(found_index)
        matched_rules.add(norm_value)
        selected.append((found_index, norm_value))

    if not selected:
        raise ValueError("Nenhuma linha da base corresponde aos NSEQs informados.")

    selected.sort(key=lambda item: (order_map.get(item[1], float("inf")), item[0]))
    ordered_indices = [idx for idx, _ in selected]
    return df.loc[ordered_indices].reset_index(drop=True)

def apply_excel_formatting(path: Path, sheet_name: str) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return
    ws = wb[sheet_name]
    if ws.max_row < 1 or ws.max_column < 1:
        wb.save(path)
        return

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "E2"
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    odd_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    even_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="D0D0D0")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for row_idx in range(2, ws.max_row + 1):
        fill = odd_fill if row_idx % 2 == 0 else even_fill
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)

# ====== NÚCLEO ADAPTADO PARA API ======
def run(nseq_list: list[str], negociacao_file_path: str | Path, modelo_file_path: str | Path | None = None) -> Path:
    """
    Função principal adaptada para receber dados da API.
    """
    if not nseq_list:
        raise ValueError("Nenhum NSEQ fornecido para processamento.")

    # 1. Definir o arquivo de modelo (Spec)
    if modelo_file_path and Path(modelo_file_path).exists():
        spec_path = Path(modelo_file_path)
    else:
        # Tenta achar pelo padrão caso o usuário não tenha enviado
        candidates = sorted(MODELOS_DIR.glob("CASAS_BAHIA_*report*.xlsx"))
        if candidates:
            spec_path = candidates[0]
        else:
            # Fallback final
            spec_path = MODELOS_DIR / "CASAS_BAHIA_21102025_REPORT - Criação report.xlsx"
            if not spec_path.exists():
                raise FileNotFoundError("Planilha de modelo da Casas Bahia não encontrada.")

    # 2. Criar objetos Rule a partir da lista de NSEQs
    rules = []
    for nseq in nseq_list:
        norm = normalize_value(nseq)
        if norm:
            rules.append(Rule(raw=nseq, normalized=norm, columns=None))
            
    if not rules:
        raise ValueError("Nenhum NSEQ válido após normalização.")

    # 3. Carregar especificações do modelo
    header_map, final_order, ref_df = load_spec(spec_path)
    base_report_df = load_reference_report(spec_path)

    # 4. Carregar Planilha de Renegociação enviada
    input_path = Path(negociacao_file_path)
    if not input_path.exists():
        raise FileNotFoundError("Arquivo de RelNegociacao não encontrado.")

    if input_path.suffix.lower() in {".xlsx", ".xls"}:
        input_df = pd.read_excel(input_path, sheet_name=0)
    else:
        input_df = pd.read_csv(input_path, sep=";", encoding="utf-8", engine="python")
        if input_df.shape[1] == 1:
            input_df = pd.read_csv(input_path, sep=",", encoding="utf-8", engine="python")

    input_df = normalize_columns(input_df)

    # 5. Filtrar e Construir o Report
    filtered_df = filter_by_rules(input_df, rules)
    report_df = build_report(header_map, final_order, filtered_df, base_report_df, rules, ref_df)
    report_df = report_df.drop(columns=["NSEQ"], errors="ignore")

    # 6. Salvar e Formatar
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    agora = datetime.now()
    out_path = REPORT_DIR / f"Report_Casas_Bahia_{agora.strftime('%d%m%Y_%H%M%S')}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)

    apply_excel_formatting(out_path, OUTPUT_SHEET_NAME)

    return out_path

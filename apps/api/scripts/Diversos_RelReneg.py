#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Diversos_RelReneg.py adaptado para rodar como API no Django/Railway (Em Memória).
"""

import io
import os
import sys
import unicodedata
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import numbers
from pathlib import Path

# ====== CONFIGURAÇÕES DE CAMINHOS PARA SERVIDOR ======
BASE_DIR = Path(__file__).resolve().parent.parent
MODELOS_DIR = BASE_DIR / "modelos"

# colunas que queremos no output (ajuste se necessário)
COLUNAS_DESEJADAS = [
    'NSEQ','ONDA','NOME LOTE','DATA LOTE CLIENTE','DATA LOTE INVEST','BANDEIRA','Empresa',
    'DENOMINACAO/ NOME','CONTRATO','CENTRO DE CUSTO','Solicitante','ENDEREÇO CLIENTE',
    'INICIO CONTRATO','TERMINO CONTRATO','ALUGUEL DEVIDO','DATA PRÓX. REAJUSTE','INDICE',
    'Valor CTO','M² AREA TERRENO','M² AREA CONSTRUIDA','M² AREA TOTAL','DADOS LOCADOR (A)',
    'DADOS FAVORECIDO (A)','Data Início Negociação','Data Fim Negociação','Data Solicitação Minuta',
    'Data Recebimento Minuta','Data Envio Locador','Data Entrega Formalizada','Data Conclusão',
    'Data Cancelamento','Motivo Cancelamento','Proposta','Contra Proposta','Motivadores sem Exito',
    'Descrição Motivadores sem Exito','Alterou Mês Reajuste','Nova Data de Reajuste','RENOVACAO?',
    'NOVO FIM DE VIGENCIA','PRAZO RENOVADO','RETROATIVO','VALOR DEB. DE RETROATIVOS','VALOR NEGOCIADO',
    'Indicador - da Sub. De ind','SUBSTITUICAO INDICE BASE','NOVO INDICE DE REAJUSTE',
    'Indicador - Isenção do Indice','ISENCAO INDICE','% Desconto Índice','REDUÇÃO / DESCONTO / MAJORAÇÃO',
    'VALOR - RED / DES / MAJ','% Red / Des / Maj','Início Captura','Fim Captura','Distrato','Data Distrato',
    'valor DRS','Data início(DRS)','Data fim(DRS)','Alteração Cadastral',
    'Economia até  12 meses - Substituição Índice','Economia até  12 meses Desconto/Redução',
    'Economia até  12 meses - Isenção Indice','Economia até  12 meses - Limitador do índice',
    'Economia Fim do Contrato - Substituição Índice','Economia Fim do Contrato -  Desconto/Redução',
    'Economia Fim do Contrato -  Isenção Indice','Economia Fim do Contrato -  Limitador do índice',
    'Status','Situação','Negociador','Resp. Adm.','Data Historico','Ultimo Historico'
]

DATE_FORMAT = "%d-%m-%Y"

# ------------------------
# UTILITÁRIOS
# ------------------------
def normalize_text(s: str) -> str:
    s = "" if s is None else str(s)
    nfkd = unicodedata.normalize("NFKD", s)
    only_ascii = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
    return only_ascii.upper().strip()

# ------------------------
# ESCRITA SEGURA (EM MEMÓRIA)
# ------------------------
def salvar_df_no_modelo_com_table_buffer(df_salvar: pd.DataFrame, modelo_file, remove_report: bool = False) -> io.BytesIO:
    """
    Injeta o DataFrame na aba 'Base' do modelo fornecido (em memória),
    formata como Tabela Oficial do Excel e retorna um novo buffer.
    """
    output_buffer = io.BytesIO()
    
    try:
        # Se houver modelo, carrega dele
        if modelo_file:
            if hasattr(modelo_file, 'seek'):
                modelo_file.seek(0)
            wb = load_workbook(modelo_file)
            
            if remove_report and "Report" in wb.sheetnames:
                del wb["Report"]
                
            if "Base" not in wb.sheetnames:
                ws = wb.create_sheet("Base")
            else:
                ws = wb["Base"]
                # Limpa a aba Base existente
                if ws.max_row > 0:
                    ws.delete_rows(1, ws.max_row)
        else:
            # Se não houver modelo (fallback extremo), cria um do zero
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Base"
            if not remove_report:
                ws_report = wb.create_sheet("Report")
                ws_report.cell(row=1, column=1, value="Coloque aqui seu modelo de Report que puxa da Base")

        # Escreve os cabeçalhos
        for col_idx, col_name in enumerate(df_salvar.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Função interna para converter tipos do Pandas para tipos nativos do Python/Excel
        def _to_python_value(v):
            try:
                if pd.isna(v):
                    return None
            except Exception:
                pass
            try:
                if isinstance(v, pd.Timestamp):
                    return v.to_pydatetime()
            except Exception:
                pass
            if isinstance(v, (numbers.Integral, numbers.Real, int, float)):
                return v
            if isinstance(v, bool):
                return v
            return v

        # Escreve os dados linha por linha
        for r_idx, row in enumerate(df_salvar.itertuples(index=False, name=None), start=2):
            for c_idx, cell_value in enumerate(row, start=1):
                val = _to_python_value(cell_value)
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Formata como Tabela Oficial do Excel
        nrows = df_salvar.shape[0] + 1
        ncols = df_salvar.shape[1] if df_salvar.shape[1] > 0 else 1
        last_col = get_column_letter(ncols)
        ref = f"A1:{last_col}{nrows}"

        existing_table_name = None
        if hasattr(ws, "_tables") and ws._tables:
            try:
                existing_table_name = ws._tables[0].displayName
            except Exception:
                existing_table_name = None
            ws._tables = []

        table_name = existing_table_name or "BaseTable"
        table_name = table_name.replace(" ", "_")
        
        tbl = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

        # Salva o resultado no buffer de saída
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    except Exception as e:
        raise RuntimeError(f"Erro ao injetar dados no modelo Excel: {str(e)}")

# =========================
# NÚCLEO ADAPTADO PARA API
# =========================
def processar_relatorio_diversos(arquivo_excel_em_memoria, nseq_string: str, arquivo_modelo_em_memoria=None) -> io.BytesIO:
    """
    Recebe o arquivo em memória, a string de NSEQs e o modelo opcional.
    Filtra os dados e injeta no modelo Excel, retornando tudo em memória.
    """
    
    # 1. Tratar a string de NSEQs
    nseq_list = [n.strip() for n in nseq_string.split(',') if n.strip()]
    if not nseq_list:
        raise ValueError("Nenhum NSEQ fornecido para processamento.")

    # Limpa os NSEQs recebidos para garantir que sejam inteiros
    nseqs_int = set()
    for n in nseq_list:
        try:
            nseqs_int.add(int(float(str(n).strip())))
        except ValueError:
            continue

    if not nseqs_int:
        raise ValueError("Nenhum NSEQ numérico válido fornecido.")

    # 2. Carrega a base de renegociação
    if hasattr(arquivo_excel_em_memoria, 'seek'):
        arquivo_excel_em_memoria.seek(0)
        
    df = pd.read_excel(arquivo_excel_em_memoria, engine="openpyxl")

    if 'BANDEIRA' not in df.columns or 'NSEQ' not in df.columns:
        raise ValueError("A planilha de renegociação deve conter as colunas 'BANDEIRA' e 'NSEQ'.")

    # 3. Prepara df para filtros
    df = df.copy()
    df['_BANDEIRA_NORM'] = df['BANDEIRA'].astype(str).apply(normalize_text)
    df['_NSEQ_INT'] = pd.to_numeric(df['NSEQ'], errors='coerce').astype('Int64')

    # 4. Filtra pelos NSEQs fornecidos
    df_filtrado = df[df['_NSEQ_INT'].isin(nseqs_int)]

    if df_filtrado.empty:
        raise ValueError("Nenhum registro encontrado na planilha para os NSEQs informados.")

    # Seleciona apenas as colunas desejadas que existem no DataFrame
    df_salvar = df_filtrado[[c for c in COLUNAS_DESEJADAS if c in df_filtrado.columns]].copy()

    # 5. Lógica de definição do modelo
    modelo_file = None
    if arquivo_modelo_em_memoria:
        modelo_file = arquivo_modelo_em_memoria
    else:
        # Tenta buscar no servidor se o usuário não enviou
        reference_path = MODELOS_DIR / "Report_RelReneg.xlsx"
        if reference_path.exists():
            # Lê o arquivo do disco para a memória para padronizar o fluxo
            with open(reference_path, "rb") as f:
                modelo_file = io.BytesIO(f.read())
        else:
            # Se não achar, o código vai criar um arquivo do zero (fallback)
            modelo_file = None

    # 6. Salva o arquivo em memória usando a função que preserva o modelo
    output_formatado = salvar_df_no_modelo_com_table_buffer(
        df_salvar=df_salvar,
        modelo_file=modelo_file,
        remove_report=False
    )

    return output_formatado
